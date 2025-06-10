#!/usr/bin/env python3
"""
Direct PRE File Parser - Excel to IndustrialQuotation Model
Converts Excel files with PRE format directly to IndustrialQuotation objects
"""

import logging
import decimal
from typing import List, Optional, Any, Dict
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from openpyxl import load_workbook

# Import unified models
import sys
import os
sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from models.quotation_models import (
    IndustrialQuotation, ProjectInfo, ProjectParameters, SalesInfo,
    ProductGroup, QuotationCategory, QuotationItem, QuotationTotals,
    CurrencyType, CategoryType, ParserType
)

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# =============================================================================
# CONSTANTS
# =============================================================================

# Excel Column Indices (1-based)
class ExcelColumns:
    COD = 1
    CODICE = 3
    DENOMINAZIONE = 4
    QTA = 5
    LIST_UNIT = 6
    LISTINO = 7
    SUB_TOT_LISTINO = 8
    SUB_TOT_CODICE = 9
    TOTALE = 10
    GRUPPI = 11
    TOTALE_OFFERTA = 12
    VALUTA = 13
    TOTALE_CODICE = 14
    TOTALE_SCONTATO = 15
    NOTE = 16
    COD_LISTINO = 17
    TOTAL_DISCOUNTED = 18
    COSTO_UNITARIO = 19
    COSTO = 20
    SUB_TOT_COSTO = 21
    TOTALE_COSTO = 22

# Excel Row Constants
class ExcelRows:
    HEADER_ROW = 17
    DATA_START_ROW = 18

# Project Information Cell Positions (row, column)
class ProjectInfoCells:
    PROJECT_ID = (1, 1)  # Row 1, Column A
    CUSTOMER = (3, 7)    # Row 3, Column G
    DOC_PERCENTAGE = (8, 2)     # Row 8, Column B
    PM_PERCENTAGE = (9, 2)      # Row 9, Column B
    FINANCIAL_COSTS = (10, 2)   # Row 10, Column B
    CURRENCY = (11, 2)          # Row 11, Column B
    EXCHANGE_RATE = (12, 2)     # Row 12, Column B
    WASTE_DISPOSAL = (13, 2)    # Row 13, Column B
    WARRANTY_PERCENTAGE = (8, 11) # Row 8, Column K

# Identification Patterns
class IdentificationPatterns:
    GROUP_PREFIX = 'TXT-'
    CATEGORY_CODE_LENGTH = 4
    HEADER_CODE = 'COD'
    MDC_SHEET_PREFIX = 'MDC'

class MDCRows:
    DATA_START_ROW = 15              # Data starts from row 15
    HEADER_ROW = 15                  # Headers are in row 15
    
class MDCColumns:
    COD = 1
    DESCRIPTION = 2
    DIRECT_COST_EUR = 4
    PRICELIST_EUR = 5
    OFFER_EUR = 6
    SALE_EUR = 7
    COMMISSION_COST_EUR = 9
    FINANCIAL_FEE_EUR = 10
    PROJECT_MANAGEMENT_COST_EUR = 11
    WARRANTY_FEE_EUR = 12
    H24_FIRST_YEAR_COST_EUR = 13
    WASTE_DISPOSAL_COST_EUR = 14
    DOCUMENTATION_COST_EUR = 15
    MARGIN_EUR = 16
    MARGIN_PERCENTAGE = 17
    

# =============================================================================
# DIRECT PRE FILE PARSER
# =============================================================================

class DirectPreFileParser:
    """Direct parser for converting PRE Excel files to IndustrialQuotation objects"""
    
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.workbook = None
        self.ws = None
        
    def load_workbook(self) -> None:
        """Load the Excel workbook"""
        try:
            self.workbook = load_workbook(str(self.file_path), data_only=True)
            self.ws = self.workbook['OFFER1']
            logger.info(f"Loaded workbook with {self.ws.max_row} rows and {self.ws.max_column} columns")
        except FileNotFoundError:
            raise FileNotFoundError(f"File not found: {self.file_path}")
        except Exception as e:
            raise Exception(f"Unable to load workbook: {str(e)}")
    
    def extract_project_info(self) -> ProjectInfo:
        """Extract project information and create ProjectInfo object"""
        # Extract basic project info
        project_id_cell = self.ws.cell(row=ProjectInfoCells.PROJECT_ID[0], column=ProjectInfoCells.PROJECT_ID[1])
        customer_cell = self.ws.cell(row=ProjectInfoCells.CUSTOMER[0], column=ProjectInfoCells.CUSTOMER[1])
        
        project_id = self._extract_after_colon(project_id_cell.value, "")
        customer = self._extract_after_colon(customer_cell.value, "")
        
        # Extract parameters
        doc_cell = self.ws.cell(row=ProjectInfoCells.DOC_PERCENTAGE[0], column=ProjectInfoCells.DOC_PERCENTAGE[1])
        pm_cell = self.ws.cell(row=ProjectInfoCells.PM_PERCENTAGE[0], column=ProjectInfoCells.PM_PERCENTAGE[1])
        financial_cell = self.ws.cell(row=ProjectInfoCells.FINANCIAL_COSTS[0], column=ProjectInfoCells.FINANCIAL_COSTS[1])
        currency_cell = self.ws.cell(row=ProjectInfoCells.CURRENCY[0], column=ProjectInfoCells.CURRENCY[1])
        exchange_cell = self.ws.cell(row=ProjectInfoCells.EXCHANGE_RATE[0], column=ProjectInfoCells.EXCHANGE_RATE[1])
        waste_cell = self.ws.cell(row=ProjectInfoCells.WASTE_DISPOSAL[0], column=ProjectInfoCells.WASTE_DISPOSAL[1])
        warranty_cell = self.ws.cell(row=ProjectInfoCells.WARRANTY_PERCENTAGE[0], column=ProjectInfoCells.WARRANTY_PERCENTAGE[1])
        
        # Create parameters object
        parameters = ProjectParameters(
            doc_percentage=self._safe_decimal(self._extract_after_colon(doc_cell.value)),
            pm_percentage=self._safe_decimal(self._extract_after_colon(pm_cell.value)),
            financial_costs=self._safe_decimal(self._extract_after_colon(financial_cell.value)),
            currency=self._extract_after_colon(currency_cell.value, "EUR"),
            exchange_rate=self._safe_decimal(self._extract_after_colon(exchange_cell.value), Decimal("1.0")),
            waste_disposal=self._safe_decimal(self._extract_after_colon(waste_cell.value)),
            warranty_percentage=self._safe_decimal(self._extract_after_colon(warranty_cell.value)),
            is_24h_service=False  # Default value, could be extracted if present
        )
        
        # Create sales info object (defaults for PRE files)
        sales_info = SalesInfo(
            area_manager=None,
            agent=None,
            commission_percentage=Decimal("0.0"),
            author=None
        )
        
        return ProjectInfo(
            id=project_id,
            customer=customer,
            parameters=parameters,
            sales_info=sales_info
        )
    
    def extract_product_groups(self) -> List[ProductGroup]:
        """Extract product groups, categories, and items directly as model objects"""
        product_groups = []
        current_group = None
        current_category = None
        
        # Extract MDC data
        self.mdc_data = self.extract_mdc_offer_data()
        
        # Start from data start row
        for row in range(ExcelRows.DATA_START_ROW, self.ws.max_row + 1):
            # Get cell values
            cod_val = self.ws.cell(row=row, column=ExcelColumns.COD).value
            codice_val = self.ws.cell(row=row, column=ExcelColumns.CODICE).value
            denominazione_val = self.ws.cell(row=row, column=ExcelColumns.DENOMINAZIONE).value
            qta_val = self.ws.cell(row=row, column=ExcelColumns.QTA).value
            listino_val = self.ws.cell(row=row, column=ExcelColumns.LIST_UNIT).value
            listino_tot_val = self.ws.cell(row=row, column=ExcelColumns.LISTINO).value
            sub_tot_listino_val = self.ws.cell(row=row, column=ExcelColumns.SUB_TOT_LISTINO).value
            sub_tot_codice_val = self.ws.cell(row=row, column=ExcelColumns.SUB_TOT_CODICE).value
            tot_val = self.ws.cell(row=row, column=ExcelColumns.TOTALE).value
            gruppi_val = self.ws.cell(row=row, column=ExcelColumns.GRUPPI).value
            tot_offer_val = self.ws.cell(row=row, column=ExcelColumns.TOTALE_OFFERTA).value
            note_val = self.ws.cell(row=row, column=ExcelColumns.NOTE).value
            cod_listino_val = self.ws.cell(row=row, column=ExcelColumns.COD_LISTINO).value
            costo_unitario_val = self.ws.cell(row=row, column=ExcelColumns.COSTO_UNITARIO).value
            costo_val = self.ws.cell(row=row, column=ExcelColumns.COSTO).value
            sub_tot_costo_val = self.ws.cell(row=row, column=ExcelColumns.SUB_TOT_COSTO).value
            tot_costo_val = self.ws.cell(row=row, column=ExcelColumns.TOTALE_COSTO).value
            
            # Check if this is a group header
            if codice_val and str(codice_val).startswith(IdentificationPatterns.GROUP_PREFIX):
                # Save previous group if exists
                if current_group:
                    product_groups.append(current_group)
                
                # Start new group
                current_group = ProductGroup(
                    group_id=str(codice_val),
                    group_name=str(denominazione_val) if denominazione_val else "",
                    quantity=self._safe_int(qta_val, 1),
                    categories=[]
                )
                current_category = None
                logger.info(f"Found group: {codice_val}")
                
            # Check if this is a category
            elif cod_val and len(str(cod_val).strip()) == IdentificationPatterns.CATEGORY_CODE_LENGTH and current_group:
                category_type = self._determine_category_type(str(cod_val))
                
                current_category = QuotationCategory(
                    category_id=str(cod_val),
                    category_name=str(denominazione_val) if denominazione_val else "",
                    wbe=codice_val,
                    items=[],
                    pricelist_subtotal=float(self._safe_decimal(sub_tot_listino_val)),
                    cost_subtotal=float(self._safe_decimal(sub_tot_costo_val)),
                    total_cost=float(self._safe_decimal(tot_costo_val)),
                    #offer_price=float(self._safe_decimal(tot_offer_val)) if tot_offer_val else None,
                    offer_price = (
                        float(self._safe_decimal(
                            self.mdc_data.get(str(cod_val) + '_' + str(current_group.group_name) + '_' + str(tot_offer_val), {}).get(MDCColumns.SALE_EUR)
                        ))
                        if self.mdc_data.get(str(cod_val) + '_' + str(current_group.group_name) + '_' + str(tot_offer_val), {}).get(MDCColumns.SALE_EUR) is not None
                        else None
                    ),
                    margin_amount = (
                        float(self._safe_decimal(
                            self.mdc_data.get(str(cod_val) + '_' + str(current_group.group_name) + '_' + str(tot_offer_val), {}).get(MDCColumns.MARGIN_EUR)
                        ))
                        if self.mdc_data.get(str(cod_val) + '_' + str(current_group.group_name) + '_' + str(tot_offer_val), {}).get(MDCColumns.MARGIN_EUR) is not None
                        else None
                    ),
                    margin_percentage = (
                        float(self._safe_decimal(
                            self.mdc_data.get(str(cod_val) + '_' + str(current_group.group_name) + '_' + str(tot_offer_val), {}).get(MDCColumns.MARGIN_PERCENTAGE)
                        ))
                        if self.mdc_data.get(str(cod_val) + '_' + str(current_group.group_name) + '_' + str(tot_offer_val), {}).get(MDCColumns.MARGIN_PERCENTAGE) is not None
                        else None
                    )
                )
                current_group.categories.append(current_category)
                logger.info(f"Found category: {cod_val}")
                
            # Check if this is an item
            elif (codice_val and denominazione_val and current_category 
                  and not str(codice_val).startswith(IdentificationPatterns.GROUP_PREFIX) 
                  and not str(codice_val).startswith(IdentificationPatterns.HEADER_CODE)):
                
                item = QuotationItem(
                    position=str(row),
                    code=str(codice_val),
                    description=str(denominazione_val),
                    quantity=float(self._safe_decimal(qta_val)),
                    pricelist_unit_price=float(self._safe_decimal(listino_val)),
                    pricelist_total_price=float(self._safe_decimal(listino_tot_val)),
                    unit_cost=float(self._safe_decimal(costo_unitario_val)),
                    total_cost=float(self._safe_decimal(costo_val))
                )
                
                current_category.items.append(item)
                logger.debug(f"Found item: {codice_val}")
        
        # Add the last group if exists
        if current_group:
            product_groups.append(current_group)
            
        # add H24 from MDC data
        # Find H24 MDC data (key starts with '_H24 PRIMO ANNO')
        h24_key = next((k for k in self.mdc_data if k.startswith('A2ZZ_H24 PRIMO ANNO')), None)
        if h24_key:
            h24_data = self.mdc_data[h24_key]
        # If H24 MDC data is found, create a group and category for it and add to product_groups
        if h24_key and h24_data:
            # Create a QuotationItem for H24 if relevant fields exist
            h24_item = QuotationItem(
                position="H24",
                code=str(h24_data.get(MDCColumns.COD, "H24")),
                description=str(h24_data.get(MDCColumns.DESCRIPTION, "H24 PRIMO ANNO")),
                quantity=1.0,
                pricelist_unit_price=float(self._safe_decimal(h24_data.get(MDCColumns.PRICELIST_EUR, 0.0))),
                pricelist_total_price=float(self._safe_decimal(h24_data.get(MDCColumns.PRICELIST_EUR, 0.0))),
                unit_cost=float(self._safe_decimal(h24_data.get(MDCColumns.DIRECT_COST_EUR, 0.0))),
                total_cost=float(self._safe_decimal(h24_data.get(MDCColumns.DIRECT_COST_EUR, 0.0)))
            )

            # Create a QuotationCategory for H24
            h24_category = QuotationCategory(
                category_id="H24",
                category_code="H24",
                category_name=str(h24_data.get(MDCColumns.DESCRIPTION, "H24 PRIMO ANNO")),
                wbe="",
                pricelist_subtotal=float(self._safe_decimal(h24_data.get(MDCColumns.PRICELIST_EUR, 0.0))),
                cost_subtotal=float(self._safe_decimal(h24_data.get(MDCColumns.DIRECT_COST_EUR, 0.0))),
                total_cost=float(self._safe_decimal(h24_data.get(MDCColumns.DIRECT_COST_EUR, 0.0))),
                offer_price=float(self._safe_decimal(h24_data.get(MDCColumns.SALE_EUR, 0.0))) if h24_data.get(MDCColumns.SALE_EUR) is not None else None,
                margin_amount=float(self._safe_decimal(h24_data.get(MDCColumns.MARGIN_EUR, 0.0))) if h24_data.get(MDCColumns.MARGIN_EUR) is not None else None,
                margin_percentage=float(self._safe_decimal(h24_data.get(MDCColumns.MARGIN_PERCENTAGE, 0.0))) if h24_data.get(MDCColumns.MARGIN_PERCENTAGE) is not None else None,
                items=[h24_item]
            )

            # Create a ProductGroup for H24
            h24_group = ProductGroup(
                group_id="H24",
                group_name="H24 PRIMO ANNO",
                quantity=1,
                categories=[h24_category],
                total_cost_value=float(self._safe_decimal(h24_data.get(MDCColumns.DIRECT_COST_EUR, 0.0))),
                total_offer_value=float(self._safe_decimal(h24_data.get(MDCColumns.OFFER_EUR, 0.0))) if h24_data.get(MDCColumns.OFFER_EUR) is not None else None,
                total_pricelist_value=float(self._safe_decimal(h24_data.get(MDCColumns.PRICELIST_EUR, 0.0))),
                item_count=1
            )

            product_groups.append(h24_group)
        
        return product_groups
    
    def _find_latest_mdc_sheet(self) -> Optional[str]:
        """Find the latest MDC sheet in the workbook"""
        mdc_sheets = [name for name in self.workbook.sheetnames 
                       if name.startswith(IdentificationPatterns.MDC_SHEET_PREFIX)]
        
        if not mdc_sheets:
            return None
        
        # Sort sheets to get the latest one
        mdc_sheets.sort(reverse=True)
        return mdc_sheets[0]
    
    def extract_mdc_offer_data(self) -> Dict[str, float]:
        """Extract additional info from MDC sheet if available"""
        mdc_sheet_name = self._find_latest_mdc_sheet()
        if not mdc_sheet_name:
            logger.info("No MDC sheet found")
            return {}
        
        try:
            mdc_ws = self.workbook[mdc_sheet_name]
            logger.info(f"Processing MDC sheet: {mdc_sheet_name}")
            
            group = ''
            mdc_data = {}
            for row in range(MDCRows.DATA_START_ROW, mdc_ws.max_row + 1):
                
                cod = mdc_ws.cell(row=row, column=MDCColumns.COD).value
                description = mdc_ws.cell(row=row, column=MDCColumns.DESCRIPTION).value
                amt = mdc_ws.cell(row=row, column=MDCColumns.OFFER_EUR).value
                
                if description:
                   
                    if not cod:
                        group = description
                        cod = ''
                    
                    key = f"{cod}_{group}_{str(amt)}"
                    
                    mdc_data[key] = {
                        MDCColumns.COD: cod,
                        MDCColumns.DESCRIPTION: description,
                        MDCColumns.DIRECT_COST_EUR: mdc_ws.cell(row=row, column=MDCColumns.DIRECT_COST_EUR).value,
                        MDCColumns.PRICELIST_EUR: mdc_ws.cell(row=row, column=MDCColumns.PRICELIST_EUR).value,
                        MDCColumns.OFFER_EUR: mdc_ws.cell(row=row, column=MDCColumns.OFFER_EUR).value,
                        MDCColumns.SALE_EUR: mdc_ws.cell(row=row, column=MDCColumns.SALE_EUR).value,
                        MDCColumns.COMMISSION_COST_EUR: mdc_ws.cell(row=row, column=MDCColumns.COMMISSION_COST_EUR).value,
                        MDCColumns.FINANCIAL_FEE_EUR: mdc_ws.cell(row=row, column=MDCColumns.FINANCIAL_FEE_EUR).value,
                        MDCColumns.PROJECT_MANAGEMENT_COST_EUR: mdc_ws.cell(row=row, column=MDCColumns.PROJECT_MANAGEMENT_COST_EUR).value,
                        MDCColumns.WARRANTY_FEE_EUR: mdc_ws.cell(row=row, column=MDCColumns.WARRANTY_FEE_EUR).value,
                        MDCColumns.H24_FIRST_YEAR_COST_EUR: mdc_ws.cell(row=row, column=MDCColumns.H24_FIRST_YEAR_COST_EUR).value,
                        MDCColumns.WASTE_DISPOSAL_COST_EUR: mdc_ws.cell(row=row, column=MDCColumns.WASTE_DISPOSAL_COST_EUR).value,
                        MDCColumns.DOCUMENTATION_COST_EUR: mdc_ws.cell(row=row, column=MDCColumns.DOCUMENTATION_COST_EUR).value,
                        MDCColumns.MARGIN_EUR: mdc_ws.cell(row=row, column=MDCColumns.MARGIN_EUR).value,
                        MDCColumns.MARGIN_PERCENTAGE: mdc_ws.cell(row=row, column=MDCColumns.MARGIN_PERCENTAGE).value
                    }
                    
                    #logger.debug(f"MDC offer: {description} = €{mdc_data[key][MDCColumns.OFFER_EUR]:,.2f}")
            
            logger.info(f"Extracted {len(mdc_data)} offer prices from MDC sheet")
            return mdc_data
            
        except Exception as e:
            logger.warning(f"Error reading MDC sheet {mdc_sheet_name}: {e}")
            return {}
    
    def calculate_totals(self, product_groups: List[ProductGroup], parameters: ProjectParameters) -> QuotationTotals:
        """Calculate total costs and fees"""
        total_pricelist = Decimal("0.0")
        total_cost = Decimal("0.0")
        total_offer = Decimal("0.0")
        offer_margin = Decimal("0.0")
        offer_margin_percentage = Decimal("0.0")
        
        # Sum up costs from all categories
        for group in product_groups:
            for category in group.categories:
                # Use offer_price if available, otherwise pricelist_subtotal
                category_pricelist  = Decimal(str(category.pricelist_subtotal)) if category.pricelist_subtotal is not None else Decimal("0.0")
                category_cost = Decimal(str(category.cost_subtotal)) if category.cost_subtotal is not None else Decimal("0.0")
                category_offer = Decimal(str(category.offer_price)) if category.offer_price is not None else Decimal("0.0")
                category_margin_amount = Decimal(str(category.margin_amount)) if category.margin_amount is not None else Decimal("0.0")

                total_pricelist += category_pricelist
                total_cost += category_cost
                total_offer += category_offer
                offer_margin += category_margin_amount
                offer_margin_percentage = offer_margin / total_offer * 100
        
        return QuotationTotals(
            total_pricelist=float(self._round_decimal(total_pricelist)),
            total_cost=float(self._round_decimal(total_cost)),
            total_offer=float(self._round_decimal(total_offer)),
            offer_margin=float(self._round_decimal(offer_margin)),
            offer_margin_percentage=float(self._round_decimal(offer_margin_percentage))
        )
    
    def close(self):
        """Close the workbook and release file handles"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self.ws = None

    def parse(self) -> IndustrialQuotation:
        """Main parsing method - returns IndustrialQuotation object directly"""
        logger.info(f"Starting direct parsing of PRE file: {self.file_path}")
        
        try:
            self.load_workbook()
            
            # Extract all sections as model objects
            project_info = self.extract_project_info()
            product_groups = self.extract_product_groups()
            totals = self.calculate_totals(product_groups, project_info.parameters)
            
            # Create final IndustrialQuotation object
            quotation = IndustrialQuotation(
                project=project_info,
                product_groups=product_groups,
                totals=totals,
                source_file=str(self.file_path),
                parser_type=ParserType.PRE_FILE_PARSER
            )
            
            logger.info(f"Direct parsing completed. Found {len(product_groups)} product groups")
            return quotation
        finally:
            # Always close the workbook to release file handles
            self.close()
    
    # Helper methods
    def _safe_cell_value(self, row: int, column: int, default: Any = None) -> Any:
        """Safely extract cell value by row and column index"""
        try:
            cell_value = self.ws.cell(row=row, column=column).value
            return cell_value if cell_value is not None else default
        except Exception:
            return default
        
    def _safe_decimal(self, value: Any, default: Decimal = None) -> Decimal:
        """Safely convert value to Decimal"""
        if value is None or value == "":
            return default if default is not None else Decimal("0.0")
        
        try:
            # Convert to string and clean up
            str_value = str(value).strip()
            
            # Handle empty strings after stripping
            if not str_value:
                return default if default is not None else Decimal("0.0")
            
            # Handle common non-numeric values
            if str_value.lower() in ['n/a', 'na', 'null', 'none', '-', '']:
                return default if default is not None else Decimal("0.0")
                
            # Remove currency symbols and common formatting
            str_value = str_value.replace('€', '').replace('$', '').replace(',', '').strip()
            
            # Handle percentage notation
            if str_value.endswith('%'):
                str_value = str_value[:-1]
                return Decimal(str_value) / Decimal("100")
            
            return Decimal(str_value)
            
        except (ValueError, TypeError, decimal.InvalidOperation) as e:
            logger.debug(f"Could not convert '{value}' to Decimal: {e}")
            return default if default is not None else Decimal("0.0")
    
    def _safe_int(self, value: Any, default: int = 0) -> int:
        """Safely convert value to int"""
        if value is None:
            return default
        try:
            return int(value)
        except (ValueError, TypeError):
            return default
    
    def _round_decimal(self, value: Decimal) -> Decimal:
        """Round decimal to 2 places"""
        return value.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    
    def _extract_after_colon(self, value: Any, default: str = "") -> str:
        """Extract text after colon, or return the value as string"""
        if value is None:
            return default
        
        str_value = str(value)
        if ':' in str_value:
            return str_value.split(':', 1)[1].strip()
        return str_value.strip()
    
    def _determine_category_type(self, category_id: str) -> CategoryType:
        """Determine category type based on category ID"""
        if category_id.startswith('E'):
            return CategoryType.INSTALLATION
        else:
            # All other categories (W, V, etc.) are considered equipment
            return CategoryType.EQUIPMENT

# =============================================================================
# CONVENIENCE FUNCTIONS
# =============================================================================

def parse_pre_file_direct(file_path: str, output_path: Optional[str] = None) -> IndustrialQuotation:
    """
    Parse PRE Excel file directly to IndustrialQuotation object
    
    Args:
        file_path: Path to the Excel file
        output_path: Optional path to save JSON output
        
    Returns:
        IndustrialQuotation: Validated quotation model instance
    """
    parser = DirectPreFileParser(file_path)
    try:
        quotation = parser.parse()
        
        if output_path:
            quotation.save_json(output_path)
            logger.info(f"JSON output saved to {output_path}")
        
        return quotation
    finally:
        parser.close()

def validate_pre_file(file_path: str) -> Dict[str, Any]:
    """
    Validate PRE file and return validation results
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        Dictionary with validation results
    """
    parser = DirectPreFileParser(file_path)
    try:
        quotation = parser.parse()
        
        # Run validation checks
        validation_results = quotation.validate_totals_consistency()
        summary_stats = quotation.get_summary_stats()
        
        return {
            "is_valid": True,
            "validation_results": validation_results,
            "summary_stats": summary_stats,
            "errors": []
        }
        
    except Exception as e:
        return {
            "is_valid": False,
            "validation_results": {},
            "summary_stats": {},
            "errors": [str(e)]
        }
    finally:
        parser.close()

if __name__ == "__main__":
    # Example usage
    input_file = "input/test_pre.xlsm"
    output_file = "output/quotation_direct.json"
    
    try:
        quotation = parse_pre_file_direct(input_file, output_file)
        stats = quotation.get_summary_stats()
        
        print("✅ Direct parsing completed successfully!")
        print(f"📋 Project ID: {stats['project_id']}")
        print(f"🏢 Customer: {quotation.project.customer}")
        print(f"📦 Product groups: {stats['total_groups']}")
        print(f"📂 Categories: {stats['total_categories']}")
        print(f"📄 Items: {stats['total_items']}")
        print(f"💵 Currency: {quotation.project.parameters.currency}")
        print(f"💰 Grand total pricelist: €{quotation.totals.total_pricelist:,.2f}")
        print(f"💰 Grand total cost: €{quotation.totals.total_cost:,.2f}")
        print(f"💰 Grand total offer: €{quotation.totals.total_offer:,.2f}")
        print(f"💰 Grand total offer margin: €{quotation.totals.offer_margin:,.2f}")
        print(f"💰 Grand total offer margin percentage: {quotation.totals.offer_margin_percentage:.2f}%")
        
        # Validate consistency
        validation = quotation.validate_totals_consistency()
        print(f"\n🔍 Validation results:")
        for check, is_valid in validation.items():
            status = "✅" if is_valid else "❌"
            print(f"   {status} {check}")
        
    except Exception as e:
        logger.error(f"Error parsing Excel file: {e}")
        raise 