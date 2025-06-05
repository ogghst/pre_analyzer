"""
Analisi Profittabilita Excel to JSON Parser
Converts Excel files with profitability analysis format to structured JSON according to schema
"""

import json
import logging
from typing import Dict, List, Optional, Any
import pandas as pd
from openpyxl import load_workbook

# Import unified models and field mappings
import sys
import os
sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from models import IndustrialQuotation, FieldMapper

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# =============================================================================
# CONSTANTS
# =============================================================================

# Excel Column Indices (1-based) - Complete mapping for all 81 columns
class ExcelColumns:
    # Basic project columns (1-21)
    COD = 1                          # COD
    PRIORITY_ORDER = 2               # Priorità per ordinamento
    PRIORITY = 3                     # Priorità
    LINE_NUMBER = 4                  # NUMERA LE RIGHE
    WBS = 5                          # WBS
    WBE = 6                          # WBE
    POSITION = 7                     # POS.
    CODICE = 8                       # CODICE
    COD_LISTINO = 9                  # COD.LISTINO
    DENOMINAZIONE = 10               # DENOMINAZIONE
    QTA = 11                         # QTA'
    SUB_TOT_LISTINO = 12             # SUB TOT. LISTINO
    LIST_UNIT = 13                   # LIST.UNIT.
    LISTINO_TOTALE = 14              # LISTINO TOTALE
    SUBTOT_COSTO = 15                # SUBTOT COSTO
    COSTO_UNITARIO = 16              # COSTO UNITARIO
    COSTO_TOTALE = 17                # COSTO TOTALE
    # Column 18 is empty
    COD_2 = 19                       # COD (second occurrence)
    # Column 20 is empty
    TOTALE = 21                      # TOTALE
    
    # Material and UTM columns (22-30)
    MAT = 22                         # MAT
    UTM_ROBOT = 23                   # UTM-ROBOT
    UTM_ROBOT_H = 24                 # UTM-ROBOT-H
    UTM_LGV = 25                     # UTM-LGV
    UTM_LGV_H = 26                   # UTM-LGV-H
    UTM_INTRA = 27                   # UTM-INTRA
    UTM_INTRA_H = 28                 # UTM-INTRA-H
    UTM_LAYOUT = 29                  # UTM-LAYOUT
    UTM_LAYOUT_H = 30                # UTM-LAYOUT-H
    
    # Engineering columns (31-40)
    UTE = 31                         # UTE
    UTE_H = 32                       # UTE-H
    BA = 33                          # BA
    BA_H = 34                        # BA-H
    SW_PC = 35                       # SW-PC
    SW_PC_H = 36                     # SW-PC-H
    SW_PLC = 37                      # SW-PLC
    SW_PLC_H = 38                    # SW-PLC-H
    SW_LGV = 39                      # SW-LGV
    SW_LGV_H = 40                    # SW-LGV-H
    
    # Manufacturing columns (41-50)
    MTG_MEC = 41                     # MTG-MEC
    MTG_MEC_H = 42                   # MTG-MEC-H
    MTG_MEC_INTRA = 43               # MTG-MEC-INTRA
    MTG_MEC_INTRA_H = 44             # MTG-MEC-INTRA-H
    CAB_ELE = 45                     # CAB-ELE
    CAB_ELE_H = 46                   # CAB-ELE-H
    CAB_ELE_INTRA = 47               # CAB-ELE-INTRA
    CAB_ELE_INTRA_H = 48             # CAB-ELE-INTRA-H
    COLL_BA = 49                     # COLL-BA
    COLL_BA_H = 50                   # COLL-BA-H
    
    # Testing columns (51-60)
    COLL_PC = 51                     # COLL-PC
    COLL_PC_H = 52                   # COLL-PC-H
    COLL_PLC = 53                    # COLL-PLC
    COLL_PLC_H = 54                  # COLL-PLC-H
    COLL_LGV = 55                    # COLL-LGV
    COLL_LGV_H = 56                  # COLL-LGV-H
    PM_COST = 57                     # PM
    PM_H = 58                        # PM-H
    SPESE_PM = 59                    # SPESE-PM
    DOCUMENT = 60                    # DOCUMENT
    
    # Logistics and field columns (61-70)
    DOCUMENT_H = 61                  # DOCUMENT-H
    IMBALLO = 62                     # IMBALLO
    STOCCAGGIO = 63                  # STOCCAGGIO
    TRASPORTO = 64                   # TRASPORTO
    SITE = 65                        # SITE
    SITE_H = 66                      # SITE-H
    INSTALL = 67                     # INSTALL
    INSTALL_H = 68                   # INSTALL-H
    AV_PC = 69                       # AV-PC
    AV_PC_H = 70                     # AV-PC-H
    
    # Additional field columns (71-81)
    AV_PLC = 71                      # AV-PLC
    AV_PLC_H = 72                    # AV-PLC-H
    AV_LGV = 73                      # AV-LGV
    AV_LGV_H = 74                    # AV-LGV-H
    SPESE_FIELD = 75                 # SPESE FIELD
    SPESE_VARIE = 76                 # SPESE VARIE
    AFTER_SALES = 77                 # AFTER SALES
    PROVVIGIONI_ITALIA = 78          # PROVVIGIONI ITALIA
    PROVVIGIONI_ESTERO = 79          # PROVVIGIONI ESTERO
    TESORETTO = 80                   # TESORETTO
    MONTAGGIO_BEMA_MBE_US = 81       # MONTAGGIO BEMA MBE-US

# VA21 Sheet Constants
class VA21Columns:
    WBE_BACKUP = 3                   # Column C - WBE codes (backup when Column D is None)
    WBE = 4                          # Column D - WBE codes (primary)
    OFFER_TOTAL = 25                 # Column Y - Offer totals

class VA21Rows:
    DATA_START_ROW = 18              # Data starts from row 18
    HEADER_ROW = 18                  # Headers are in row 18

# Excel Row Constants
class ExcelRows:
    HEADER_ROW = 3
    DATA_START_ROW = 4
    PROJECT_INFO_START = 1
    PROJECT_INFO_END = 6

# Identification Patterns
class IdentificationPatterns:
    GROUP_PREFIX = 'TXT'
    CATEGORY_CODE_LENGTH = 4
    HEADER_CODE = 'COD'
    VA21_SHEET_PREFIX = 'VA21'       # Prefix for VA21 sheets
    WBE_IT_SUFFIX = '-IT'            # Italian WBE suffix in NEW_OFFER1
    WBE_US_SUFFIX = '-US'            # US WBE suffix in VA21 sheets

# Project Information Cell Positions (row, column)
class ProjectInfoCells:
    PROJECT_ID = (1, 1)        # Row 1, Column A - Project identifier
    LISTINO = (2, 1)           # Row 2, Column A - Price list info

# JSON Field Names - Extended to include all columns
class JsonFields:
    # Project fields
    PROJECT = "project"
    ID = "id"
    LISTINO = "listino"
    PARAMETERS = "parameters"
    SALES_INFO = "sales_info"
    
    # Parameter fields (if available in future versions)
    DOC_PERCENTAGE = "doc_percentage"
    PM_PERCENTAGE = "pm_percentage"
    FINANCIAL_COSTS = "financial_costs"
    CURRENCY = "currency"
    EXCHANGE_RATE = "exchange_rate"
    WASTE_DISPOSAL = "waste_disposal"
    WARRANTY_PERCENTAGE = "warranty_percentage"
    IS_24H_SERVICE = "is_24h_service"
    
    # Sales info fields
    AREA_MANAGER = "area_manager"
    AGENT = "agent"
    COMMISSION_PERCENTAGE = "commission_percentage"
    AUTHOR = "author"
    
    # Product group fields
    PRODUCT_GROUPS = "product_groups"
    GROUP_ID = "group_id"
    GROUP_NAME = "group_name"
    QUANTITY = "quantity"
    CATEGORIES = "categories"
    
    # Category fields
    CATEGORY_ID = "category_id"
    CATEGORY_NAME = "category_name"
    CATEGORY_CODE = "category_code"
    WBE = "wbe"
    ITEMS = "items"
    PRICELIST_SUBTOTAL = "pricelist_subtotal"
    COST_SUBTOTAL = "cost_subtotal"
    TOTAL_COST = "total_cost"
    OFFER_PRICE = "offer_price"          # New field for VA21 offer price
    
    # Basic item fields
    POSITION = "position"
    CODE = "code"
    COD_LISTINO = "cod_listino"
    DESCRIPTION = "description"
    QTY = "quantity"
    PRICELIST_UNIT_PRICE = "list_unit_price"
    PRICELIST_TOTAL = "listino_total"
    UNIT_COST = "unit_cost"
    TOTAL_COST = "total_cost"
    INTERNAL_CODE = "internal_code"
    PRIORITY_ORDER = "priority_order"
    PRIORITY = "priority"
    LINE_NUMBER = "line_number"
    WBS = "wbs"
    TOTAL = "totale"
    
    # Material and UTM fields
    MAT = "mat"
    UTM_ROBOT = "utm_robot"
    UTM_ROBOT_H = "utm_robot_h"
    UTM_LGV = "utm_lgv"
    UTM_LGV_H = "utm_lgv_h"
    UTM_INTRA = "utm_intra"
    UTM_INTRA_H = "utm_intra_h"
    UTM_LAYOUT = "utm_layout"
    UTM_LAYOUT_H = "utm_layout_h"
    
    # Engineering fields
    UTE = "ute"
    UTE_H = "ute_h"
    BA = "ba"
    BA_H = "ba_h"
    SW_PC = "sw_pc"
    SW_PC_H = "sw_pc_h"
    SW_PLC = "sw_plc"
    SW_PLC_H = "sw_plc_h"
    SW_LGV = "sw_lgv"
    SW_LGV_H = "sw_lgv_h"
    
    # Manufacturing fields
    MTG_MEC = "mtg_mec"
    MTG_MEC_H = "mtg_mec_h"
    MTG_MEC_INTRA = "mtg_mec_intra"
    MTG_MEC_INTRA_H = "mtg_mec_intra_h"
    CAB_ELE = "cab_ele"
    CAB_ELE_H = "cab_ele_h"
    CAB_ELE_INTRA = "cab_ele_intra"
    CAB_ELE_INTRA_H = "cab_ele_intra_h"
    COLL_BA = "coll_ba"
    COLL_BA_H = "coll_ba_h"
    
    # Testing fields
    COLL_PC = "coll_pc"
    COLL_PC_H = "coll_pc_h"
    COLL_PLC = "coll_plc"
    COLL_PLC_H = "coll_plc_h"
    COLL_LGV = "coll_lgv"
    COLL_LGV_H = "coll_lgv_h"
    PM_COST = "pm_cost"
    PM_H = "pm_h"
    SPESE_PM = "spese_pm"
    DOCUMENT = "document"
    
    # Logistics and field fields
    DOCUMENT_H = "document_h"
    IMBALLO = "imballo"
    STOCCAGGIO = "stoccaggio"
    TRASPORTO = "trasporto"
    SITE = "site"
    SITE_H = "site_h"
    INSTALL = "install"
    INSTALL_H = "install_h"
    AV_PC = "av_pc"
    AV_PC_H = "av_pc_h"
    
    # Additional field fields
    AV_PLC = "av_plc"
    AV_PLC_H = "av_plc_h"
    AV_LGV = "av_lgv"
    AV_LGV_H = "av_lgv_h"
    SPESE_FIELD = "spese_field"
    SPESE_VARIE = "spese_varie"
    AFTER_SALES = "after_sales"
    PROVVIGIONI_ITALIA = "provvigioni_italia"
    PROVVIGIONI_ESTERO = "provvigioni_estero"
    TESORETTO = "tesoretto"
    MONTAGGIO_BEMA_MBE_US = "montaggio_bema_mbe_us"
    
    # Totals fields
    TOTALS = "totals"
    EQUIPMENT_TOTAL = "equipment_total"
    INSTALLATION_TOTAL = "installation_total"
    SUBTOTAL = "subtotal"
    TOTAL_LISTINO = "total_listino"
    TOTAL_COSTO = "total_costo"
    TOTAL_OFFER = "total_offer"         # New field for total offer price from VA21
    MARGIN = "margin"
    MARGIN_PERCENTAGE = "margin_percentage"
    OFFER_MARGIN = "offer_margin"       # New field for margin vs offer price
    OFFER_MARGIN_PERCENTAGE = "offer_margin_percentage"  # New field for offer margin %

# Calculation Constants
class CalculationConstants:
    DEFAULT_QUANTITY = 1
    DEFAULT_FLOAT = 0.0
    DEFAULT_INT = 0

# Log Messages
class LogMessages:
    PARSING_START = "Starting to parse {}"
    WORKBOOK_LOADED = "Loaded workbook with {} rows and {} columns"
    GROUP_FOUND = "Found group: {}"
    CATEGORY_FOUND = "Found category: {}"
    ITEM_FOUND = "Found item: {}"
    PARSING_COMPLETED = "Parsing completed. Found {} product groups"
    JSON_SAVED = "JSON output saved to {}"

# Error Messages
class ErrorMessages:
    FILE_NOT_FOUND = "File not found: {}"
    INVALID_WORKBOOK = "Unable to load workbook: {}"
    MISSING_SHEET = "Sheet '{}' not found in workbook"
    INVALID_DATA_FORMAT = "Invalid data format in row {}"

# VA21 to NEW_OFFER1 Column Mapping
class VA21FieldMapping:
    """Mapping from VA21 column names to NEW_OFFER1 JsonFields"""
    MAPPINGS = {
        # Common SAP/VA21 column names
        'PO Item': JsonFields.POSITION,
        'Material': JsonFields.CODE,
        'Description': JsonFields.DESCRIPTION,
        'Order Quantity': JsonFields.QTY,
        'Net Value': JsonFields.PRICELIST_TOTAL,
        'Unit Price': JsonFields.PRICELIST_UNIT_PRICE,
        'WBE': JsonFields.WBE,
        'Quantity': JsonFields.QTY,
        'Item': JsonFields.POSITION,
        'Material Description': JsonFields.DESCRIPTION,
        'Short Text': JsonFields.DESCRIPTION,
        'Plant': JsonFields.CODE,
        'Vendor': JsonFields.CODE,
        'Document Currency': JsonFields.CODE,
        'Total Value': JsonFields.PRICELIST_TOTAL,
        'Amount': JsonFields.PRICELIST_TOTAL,
        'Price': JsonFields.PRICELIST_UNIT_PRICE,
        'Unit': JsonFields.QTY,
        # Add Italian column names if present
        'Materiale': JsonFields.CODE,
        'Descrizione': JsonFields.DESCRIPTION,
        'Quantità': JsonFields.QTY,
        'Prezzo': JsonFields.PRICELIST_UNIT_PRICE,
        'Valore': JsonFields.PRICELIST_TOTAL,
        # Add more mappings as needed based on actual VA21 headers
    }

# =============================================================================
# MAIN PARSER CLASS
# =============================================================================

class AnalisiProfittabilitaParser:
    """Parser for converting Analisi Profittabilita Excel files to JSON format"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.ws = None
        
    def load_workbook(self):
        """Load the Excel workbook"""
        try:
            self.workbook = load_workbook(self.file_path, data_only=True)
            # Use the first worksheet (typically 'NEW_OFFER1')
            self.ws = self.workbook['NEW_OFFER1']
            logger.info(LogMessages.WORKBOOK_LOADED.format(self.ws.max_row, self.ws.max_column))
        except FileNotFoundError:
            raise FileNotFoundError(ErrorMessages.FILE_NOT_FOUND.format(self.file_path))
        except Exception as e:
            raise Exception(ErrorMessages.INVALID_WORKBOOK.format(str(e)))
    
    def extract_project_info(self) -> Dict[str, Any]:
        """Extract project information from the Excel file"""
        # Extract basic project info
        project_id_cell = self.ws.cell(row=ProjectInfoCells.PROJECT_ID[0], column=ProjectInfoCells.PROJECT_ID[1])
        listino_cell = self.ws.cell(row=ProjectInfoCells.LISTINO[0], column=ProjectInfoCells.LISTINO[1])
        
        project_id = str(project_id_cell.value) if project_id_cell.value else ""
        listino = str(listino_cell.value) if listino_cell.value else ""
        
        return {
            JsonFields.ID: project_id,
            JsonFields.LISTINO: listino,
            JsonFields.PARAMETERS: {
                JsonFields.DOC_PERCENTAGE: CalculationConstants.DEFAULT_FLOAT,
                JsonFields.PM_PERCENTAGE: CalculationConstants.DEFAULT_FLOAT,
                JsonFields.FINANCIAL_COSTS: CalculationConstants.DEFAULT_FLOAT,
                JsonFields.CURRENCY: "EUR",  # Default currency
                JsonFields.EXCHANGE_RATE: 1.0,
                JsonFields.WASTE_DISPOSAL: CalculationConstants.DEFAULT_FLOAT,
                JsonFields.WARRANTY_PERCENTAGE: CalculationConstants.DEFAULT_FLOAT,
                JsonFields.IS_24H_SERVICE: False
            },
            JsonFields.SALES_INFO: {
                JsonFields.AREA_MANAGER: None,
                JsonFields.AGENT: None,
                JsonFields.COMMISSION_PERCENTAGE: CalculationConstants.DEFAULT_FLOAT,
                JsonFields.AUTHOR: None
            }
        }
    
    def _safe_cell_value(self, row: int, column: int, default: Any = None) -> Any:
        """
        Safely extract cell value by row and column index
        
        Args:
            row: Row number (1-based)
            column: Column number (1-based)
            default: Default value if cell is empty or error occurs
            
        Returns:
            Cell value or default
        """
        try:
            cell_value = self.ws.cell(row=row, column=column).value
            if cell_value is None:
                return default
            return cell_value
        except Exception:
            return default
    
    def _safe_cell_float(self, row: int, column: int, default: float = 0.0) -> float:
        """Safely extract float value from cell"""
        return self._safe_float(self._safe_cell_value(row, column), default)
    
    def _safe_cell_int(self, row: int, column: int, default: int = 0) -> int:
        """Safely extract integer value from cell"""
        return self._safe_int(self._safe_cell_value(row, column), default)
    
    def _safe_cell_str(self, row: int, column: int, default: str = "") -> str:
        """Safely extract string value from cell"""
        value = self._safe_cell_value(row, column, default)
        return str(value) if value is not None else default
    
    def extract_product_groups(self) -> List[Dict[str, Any]]:
        """Extract product groups, categories, and items with all columns using safe column access"""
        product_groups = []
        current_group = None
        current_category = None
        
        # Start from data start row
        for row in range(ExcelRows.DATA_START_ROW, self.ws.max_row + 1):
            
            # Skip row if no priority value
            if not self._safe_cell_value(row, ExcelColumns.PRIORITY):
                continue
            
            # Extract basic identification values using safe column access
            cod_val = self._safe_cell_value(row, ExcelColumns.COD)
            codice_val = self._safe_cell_value(row, ExcelColumns.CODICE)
            denominazione_val = self._safe_cell_value(row, ExcelColumns.DENOMINAZIONE)
            qta_val = self._safe_cell_value(row, ExcelColumns.QTA)
            wbe_val = self._safe_cell_value(row, ExcelColumns.WBE)

            # Check if this is a group header (TXT in CODICE)
            if codice_val and str(codice_val).startswith(IdentificationPatterns.GROUP_PREFIX):
                # Save previous group if exists
                if current_group:
                    product_groups.append(current_group)
                
                # Start new group
                current_group = {
                    JsonFields.GROUP_ID: str(codice_val),
                    JsonFields.GROUP_NAME: str(denominazione_val) if denominazione_val else "",
                    JsonFields.QUANTITY: self._safe_int(qta_val, CalculationConstants.DEFAULT_QUANTITY),
                    JsonFields.CATEGORIES: []
                }
                current_category = None
                logger.info(LogMessages.GROUP_FOUND.format(codice_val))
                
            # Check if this is a category (4-char code in COD column)
            elif cod_val and len(str(cod_val).strip()) == IdentificationPatterns.CATEGORY_CODE_LENGTH and current_group:
                current_category = {
                    JsonFields.CATEGORY_ID: str(cod_val),
                    JsonFields.CATEGORY_CODE: str(codice_val) if codice_val else "",
                    JsonFields.CATEGORY_NAME: str(denominazione_val) if denominazione_val else "",
                    JsonFields.WBE: str(wbe_val) if wbe_val else "",
                    JsonFields.PRICELIST_SUBTOTAL: self._safe_cell_float(row, ExcelColumns.SUB_TOT_LISTINO),
                    JsonFields.COST_SUBTOTAL: self._safe_cell_float(row, ExcelColumns.SUBTOT_COSTO),
                    JsonFields.TOTAL_COST: self._safe_cell_float(row, ExcelColumns.COSTO_TOTALE),
                    JsonFields.ITEMS: [],
                }
                current_group[JsonFields.CATEGORIES].append(current_category)
                logger.info(LogMessages.CATEGORY_FOUND.format(cod_val))
                
            # Check if this is an item
            elif denominazione_val and current_category \
                and not (codice_val and str(codice_val).startswith(IdentificationPatterns.GROUP_PREFIX)) \
                and not (cod_val and len(str(cod_val).strip()) == IdentificationPatterns.CATEGORY_CODE_LENGTH) \
                and str(denominazione_val) != "DENOMINAZIONE":  # Skip header row
                
                item = {
                    # Basic identification - using safe column access
                    JsonFields.POSITION: str(row),
                    JsonFields.CODE: self._safe_cell_str(row, ExcelColumns.CODICE),
                    JsonFields.COD_LISTINO: self._safe_cell_str(row, ExcelColumns.COD_LISTINO),
                    JsonFields.DESCRIPTION: str(denominazione_val),
                    JsonFields.QTY: self._safe_cell_float(row, ExcelColumns.QTA),
                    JsonFields.PRICELIST_UNIT_PRICE: self._safe_cell_float(row, ExcelColumns.LIST_UNIT),
                    JsonFields.PRICELIST_TOTAL: self._safe_cell_float(row, ExcelColumns.LISTINO_TOTALE),
                    JsonFields.UNIT_COST: self._safe_cell_float(row, ExcelColumns.COSTO_UNITARIO),
                    JsonFields.TOTAL_COST: self._safe_cell_float(row, ExcelColumns.COSTO_TOTALE),
                    JsonFields.INTERNAL_CODE: self._safe_cell_str(row, ExcelColumns.COD_2),
                    JsonFields.PRIORITY_ORDER: self._safe_cell_int(row, ExcelColumns.PRIORITY_ORDER),
                    JsonFields.PRIORITY: self._safe_cell_int(row, ExcelColumns.PRIORITY),
                    JsonFields.LINE_NUMBER: self._safe_cell_int(row, ExcelColumns.LINE_NUMBER),
                    JsonFields.WBS: self._safe_cell_str(row, ExcelColumns.WBS),
                    JsonFields.TOTAL: self._safe_cell_float(row, ExcelColumns.TOTALE),
                    
                    # Material and UTM fields - using safe column access
                    JsonFields.MAT: self._safe_cell_float(row, ExcelColumns.MAT),
                    JsonFields.UTM_ROBOT: self._safe_cell_float(row, ExcelColumns.UTM_ROBOT),
                    JsonFields.UTM_ROBOT_H: self._safe_cell_float(row, ExcelColumns.UTM_ROBOT_H),
                    JsonFields.UTM_LGV: self._safe_cell_float(row, ExcelColumns.UTM_LGV),
                    JsonFields.UTM_LGV_H: self._safe_cell_float(row, ExcelColumns.UTM_LGV_H),
                    JsonFields.UTM_INTRA: self._safe_cell_float(row, ExcelColumns.UTM_INTRA),
                    JsonFields.UTM_INTRA_H: self._safe_cell_float(row, ExcelColumns.UTM_INTRA_H),
                    JsonFields.UTM_LAYOUT: self._safe_cell_float(row, ExcelColumns.UTM_LAYOUT),
                    JsonFields.UTM_LAYOUT_H: self._safe_cell_float(row, ExcelColumns.UTM_LAYOUT_H),
                    
                    # Engineering fields - using safe column access
                    JsonFields.UTE: self._safe_cell_float(row, ExcelColumns.UTE),
                    JsonFields.UTE_H: self._safe_cell_float(row, ExcelColumns.UTE_H),
                    JsonFields.BA: self._safe_cell_float(row, ExcelColumns.BA),
                    JsonFields.BA_H: self._safe_cell_float(row, ExcelColumns.BA_H),
                    JsonFields.SW_PC: self._safe_cell_float(row, ExcelColumns.SW_PC),
                    JsonFields.SW_PC_H: self._safe_cell_float(row, ExcelColumns.SW_PC_H),
                    JsonFields.SW_PLC: self._safe_cell_float(row, ExcelColumns.SW_PLC),
                    JsonFields.SW_PLC_H: self._safe_cell_float(row, ExcelColumns.SW_PLC_H),
                    JsonFields.SW_LGV: self._safe_cell_float(row, ExcelColumns.SW_LGV),
                    JsonFields.SW_LGV_H: self._safe_cell_float(row, ExcelColumns.SW_LGV_H),
                    
                    # Manufacturing fields - using safe column access
                    JsonFields.MTG_MEC: self._safe_cell_float(row, ExcelColumns.MTG_MEC),
                    JsonFields.MTG_MEC_H: self._safe_cell_float(row, ExcelColumns.MTG_MEC_H),
                    JsonFields.MTG_MEC_INTRA: self._safe_cell_float(row, ExcelColumns.MTG_MEC_INTRA),
                    JsonFields.MTG_MEC_INTRA_H: self._safe_cell_float(row, ExcelColumns.MTG_MEC_INTRA_H),
                    JsonFields.CAB_ELE: self._safe_cell_float(row, ExcelColumns.CAB_ELE),
                    JsonFields.CAB_ELE_H: self._safe_cell_float(row, ExcelColumns.CAB_ELE_H),
                    JsonFields.CAB_ELE_INTRA: self._safe_cell_float(row, ExcelColumns.CAB_ELE_INTRA),
                    JsonFields.CAB_ELE_INTRA_H: self._safe_cell_float(row, ExcelColumns.CAB_ELE_INTRA_H),
                    JsonFields.COLL_BA: self._safe_cell_float(row, ExcelColumns.COLL_BA),
                    JsonFields.COLL_BA_H: self._safe_cell_float(row, ExcelColumns.COLL_BA_H),
                    
                    # Testing fields - using safe column access
                    JsonFields.COLL_PC: self._safe_cell_float(row, ExcelColumns.COLL_PC),
                    JsonFields.COLL_PC_H: self._safe_cell_float(row, ExcelColumns.COLL_PC_H),
                    JsonFields.COLL_PLC: self._safe_cell_float(row, ExcelColumns.COLL_PLC),
                    JsonFields.COLL_PLC_H: self._safe_cell_float(row, ExcelColumns.COLL_PLC_H),
                    JsonFields.COLL_LGV: self._safe_cell_float(row, ExcelColumns.COLL_LGV),
                    JsonFields.COLL_LGV_H: self._safe_cell_float(row, ExcelColumns.COLL_LGV_H),
                    JsonFields.PM_COST: self._safe_cell_float(row, ExcelColumns.PM_COST),
                    JsonFields.PM_H: self._safe_cell_float(row, ExcelColumns.PM_H),
                    JsonFields.SPESE_PM: self._safe_cell_float(row, ExcelColumns.SPESE_PM),
                    JsonFields.DOCUMENT: self._safe_cell_float(row, ExcelColumns.DOCUMENT),
                    
                    # Logistics and field fields - using safe column access
                    JsonFields.DOCUMENT_H: self._safe_cell_float(row, ExcelColumns.DOCUMENT_H),
                    JsonFields.IMBALLO: self._safe_cell_float(row, ExcelColumns.IMBALLO),
                    JsonFields.STOCCAGGIO: self._safe_cell_float(row, ExcelColumns.STOCCAGGIO),
                    JsonFields.TRASPORTO: self._safe_cell_float(row, ExcelColumns.TRASPORTO),
                    JsonFields.SITE: self._safe_cell_float(row, ExcelColumns.SITE),
                    JsonFields.SITE_H: self._safe_cell_float(row, ExcelColumns.SITE_H),
                    JsonFields.INSTALL: self._safe_cell_float(row, ExcelColumns.INSTALL),
                    JsonFields.INSTALL_H: self._safe_cell_float(row, ExcelColumns.INSTALL_H),
                    JsonFields.AV_PC: self._safe_cell_float(row, ExcelColumns.AV_PC),
                    JsonFields.AV_PC_H: self._safe_cell_float(row, ExcelColumns.AV_PC_H),
                    
                    # Additional field fields - using safe column access
                    JsonFields.AV_PLC: self._safe_cell_float(row, ExcelColumns.AV_PLC),
                    JsonFields.AV_PLC_H: self._safe_cell_float(row, ExcelColumns.AV_PLC_H),
                    JsonFields.AV_LGV: self._safe_cell_float(row, ExcelColumns.AV_LGV),
                    JsonFields.AV_LGV_H: self._safe_cell_float(row, ExcelColumns.AV_LGV_H),
                    JsonFields.SPESE_FIELD: self._safe_cell_float(row, ExcelColumns.SPESE_FIELD),
                    JsonFields.SPESE_VARIE: self._safe_cell_float(row, ExcelColumns.SPESE_VARIE),
                    JsonFields.AFTER_SALES: self._safe_cell_float(row, ExcelColumns.AFTER_SALES),
                    JsonFields.PROVVIGIONI_ITALIA: self._safe_cell_float(row, ExcelColumns.PROVVIGIONI_ITALIA),
                    JsonFields.PROVVIGIONI_ESTERO: self._safe_cell_float(row, ExcelColumns.PROVVIGIONI_ESTERO),
                    JsonFields.TESORETTO: self._safe_cell_float(row, ExcelColumns.TESORETTO),
                    JsonFields.MONTAGGIO_BEMA_MBE_US: self._safe_cell_float(row, ExcelColumns.MONTAGGIO_BEMA_MBE_US)
                }
                
                current_category[JsonFields.ITEMS].append(item)
                logger.debug(LogMessages.ITEM_FOUND.format(codice_val))
        
        # Add the last group if exists
        if current_group:
            product_groups.append(current_group)
        
        # Calculate category totals
        for group in product_groups:
            for category in group[JsonFields.CATEGORIES]:
                if not category[JsonFields.PRICELIST_SUBTOTAL]:
                    category[JsonFields.PRICELIST_SUBTOTAL] = sum(
                        item[JsonFields.PRICELIST_TOTAL] for item in category[JsonFields.ITEMS]
                    )
                if not category[JsonFields.COST_SUBTOTAL]:
                    category[JsonFields.COST_SUBTOTAL] = sum(
                        item[JsonFields.TOTAL_COST] for item in category[JsonFields.ITEMS]
                    )
                if not category[JsonFields.TOTAL_COST]:
                    category[JsonFields.TOTAL_COST] = category[JsonFields.COST_SUBTOTAL]
        
        return product_groups
    
    def calculate_totals(self, product_groups: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Calculate total costs, revenues, offer prices and margins"""
        total_listino = CalculationConstants.DEFAULT_FLOAT
        total_costo = CalculationConstants.DEFAULT_FLOAT
        total_offer = CalculationConstants.DEFAULT_FLOAT
        
        # Sum up all costs, revenues, and offers from all categories
        for group in product_groups:
            for category in group[JsonFields.CATEGORIES]:
                total_listino += category.get(JsonFields.PRICELIST_SUBTOTAL, 0)
                total_costo += category.get(JsonFields.COST_SUBTOTAL, 0)
                total_offer += category.get(JsonFields.OFFER_PRICE, 0)
        
        # Calculate margins using correct formula
        margin = total_listino - total_costo
        margin_percentage = (margin / total_listino * 100) if total_listino > 0 else 0
        
        # Calculate offer-based margins using correct formula: margin% = 1 - (cost / offer)
        offer_margin = total_offer - total_costo
        offer_margin_percentage = (1 - (total_costo / total_offer)) * 100 if total_offer > 0 else 0
        
        return {
            JsonFields.TOTAL_LISTINO: round(total_listino, 2),
            JsonFields.TOTAL_COSTO: round(total_costo, 2),
            JsonFields.TOTAL_OFFER: round(total_offer, 2),
            JsonFields.MARGIN: round(margin, 2),
            JsonFields.MARGIN_PERCENTAGE: round(margin_percentage, 2),
            JsonFields.OFFER_MARGIN: round(offer_margin, 2),
            JsonFields.OFFER_MARGIN_PERCENTAGE: round(offer_margin_percentage, 2),
            JsonFields.EQUIPMENT_TOTAL: round(total_listino, 2),  # For compatibility
            JsonFields.INSTALLATION_TOTAL: CalculationConstants.DEFAULT_FLOAT,
            JsonFields.SUBTOTAL: round(total_listino, 2)
        }
    
    def parse(self) -> Dict[str, Any]:
        """Main parsing method"""
        logger.info(LogMessages.PARSING_START.format(self.file_path))
        
        self.load_workbook()
        
        # Extract all sections
        project_info = self.extract_project_info()
        product_groups = self.extract_product_groups()
        
        # Integrate VA21 offer prices into categories
        product_groups = self.integrate_va21_offers_into_categories(product_groups)
        
        # Calculate totals (including offer-based calculations)
        totals = self.calculate_totals(product_groups)
        
        # Build final structure
        result = {
            JsonFields.PROJECT: project_info,
            JsonFields.PRODUCT_GROUPS: product_groups,
            JsonFields.TOTALS: totals
        }
        
        logger.info(LogMessages.PARSING_COMPLETED.format(len(product_groups)))
        return result
    
    def parse_to_model(self) -> IndustrialQuotation:
        """
        Parse Excel file directly to IndustrialQuotation model with validation
        
        Returns:
            IndustrialQuotation: Validated quotation model instance
        """
        logger.info(f"Parsing Analisi Profittabilita file to IndustrialQuotation model: {self.file_path}")
        
        # Get raw parser data
        raw_data = self.parse()
        
        # Convert to unified model format using field mapper
        converted_data = FieldMapper.convert_ap_parser_dict(raw_data)
        
        # Create and validate model instance
        quotation = IndustrialQuotation.from_parser_dict(
            converted_data,
            source_file=self.file_path,
            parser_type="analisi_profittabilita_parser"
        )
        
        logger.info(f"Successfully created IndustrialQuotation model with {quotation.get_summary_stats()['total_items']} items")
        return quotation
    
    def _safe_float(self, value: Any, default: float = CalculationConstants.DEFAULT_FLOAT) -> float:
        """Safely convert value to float"""
        if value is None:
            return default
        try:
            return float(value)
        except (ValueError, TypeError):
            return default
    
    def _safe_int(self, value: Any, default: int = CalculationConstants.DEFAULT_INT) -> int:
        """Safely convert value to int"""
        if value is None:
            return default
        try:
            return int(value)
        except (ValueError, TypeError):
            return default

    def _find_latest_va21_sheet(self) -> Optional[str]:
        """
        Find the latest VA21 sheet in the workbook.
        Sheets are named VA21, VA21_A01, VA21_A02, etc.
        Returns the sheet name with the highest version number.
        """
        va21_sheets = [
            sheet for sheet in self.workbook.sheetnames 
            if sheet.startswith(IdentificationPatterns.VA21_SHEET_PREFIX)
        ]
        
        if not va21_sheets:
            logger.warning("No VA21 sheets found in workbook")
            return None
        
        # Sort sheets to get the latest version
        # VA21 comes before VA21_A01, VA21_A02, etc.
        va21_sheets.sort()
        latest_sheet = va21_sheets[-1]  # Last in sorted order should be latest version
        
        logger.info(f"Found VA21 sheets: {va21_sheets}, using latest: {latest_sheet}")
        return latest_sheet

    def _convert_wbe_us_to_it(self, wbe_us: str) -> str:
        """
        Convert US WBE code (ending in -US) to IT format (ending in -IT).
        
        Args:
            wbe_us: WBE code in US format (e.g., 'CC2199-A-PCZZ01-US')
            
        Returns:
            WBE code in IT format (e.g., 'CC2199-A-PCZZ01-IT')
        """
        if not wbe_us or not isinstance(wbe_us, str):
            return ""
        
        wbe_clean = wbe_us.strip()
        if wbe_clean.endswith(IdentificationPatterns.WBE_US_SUFFIX):
            # Replace -US suffix with -IT
            wbe_it = wbe_clean[:-len(IdentificationPatterns.WBE_US_SUFFIX)] + IdentificationPatterns.WBE_IT_SUFFIX
            return wbe_it
        
        # If it doesn't end with -US, return as is
        return wbe_clean

    def extract_va21_offer_data(self) -> Dict[str, float]:
        """
        Extract offer prices from the latest VA21 sheet.
        Sum all offer prices for the same WBE code to avoid duplicates.
        
        Returns:
            Dictionary mapping WBE codes to total offer prices
        """
        latest_sheet = self._find_latest_va21_sheet()
        if not latest_sheet:
            logger.warning("No VA21 sheet found, offer prices will not be available")
            return {}
        
        try:
            va21_ws = self.workbook[latest_sheet]
            wbe_offers = {}
            processed_rows = 0
            valid_offer_rows = 0
            
            logger.info(f"Extracting offer data from sheet '{latest_sheet}' (Column D for WBE, Column Y for offers)")
            
            # Extract WBE-Offer mappings starting from data row
            for row in range(VA21Rows.DATA_START_ROW, va21_ws.max_row + 1):
                processed_rows += 1
                
                # Use direct cell access for VA21 sheet (Column D primary, Column C backup for WBE)
                wbe_cell = va21_ws.cell(row=row, column=VA21Columns.WBE)
                wbe_backup_cell = va21_ws.cell(row=row, column=VA21Columns.WBE_BACKUP)
                offer_cell = va21_ws.cell(row=row, column=VA21Columns.OFFER_TOTAL)
                
                wbe_val = wbe_cell.value
                wbe_backup_val = wbe_backup_cell.value
                offer_val = offer_cell.value
                
                # Log all rows for debugging (only first 10 and last 10 to avoid spam)
                if processed_rows <= 10 or processed_rows > va21_ws.max_row - 10:
                    if wbe_val or wbe_backup_val or offer_val:
                        logger.debug(f"Row {row}: WBE_D='{wbe_val}', WBE_C='{wbe_backup_val}', Offer={offer_val}")
                
                # Determine which WBE to use
                final_wbe = None
                if wbe_val and str(wbe_val).strip() and str(wbe_val).strip() != 'None':
                    # Use Column D (primary)
                    final_wbe = str(wbe_val).strip()
                elif wbe_backup_val and str(wbe_backup_val).strip() and str(wbe_backup_val).strip() != 'None':
                    # Use Column C (backup) and convert -US to -IT if needed
                    final_wbe = self._convert_wbe_us_to_it(str(wbe_backup_val).strip())
                
                # Only process rows with valid WBE and numeric offer values
                if (final_wbe and 
                    offer_val is not None and 
                    isinstance(offer_val, (int, float))):
                    
                    offer_clean = float(offer_val)
                    valid_offer_rows += 1
                    
                    # Sum offers for the same WBE (handle multiple entries for same WBE)
                    if final_wbe not in wbe_offers:
                        wbe_offers[final_wbe] = 0
                        logger.debug(f"Row {row}: First occurrence of WBE '{final_wbe}': €{offer_clean:,.2f}")
                    else:
                        logger.debug(f"Row {row}: Additional entry for WBE '{final_wbe}': +€{offer_clean:,.2f} (previous: €{wbe_offers[final_wbe]:,.2f})")
                    
                    wbe_offers[final_wbe] += offer_clean
                    logger.debug(f"Row {row}: WBE '{final_wbe}' total now: €{wbe_offers[final_wbe]:,.2f}")
                    
                elif (wbe_val or wbe_backup_val) and offer_val is not None:
                    # Log cases where we have data but it's not being processed
                    logger.debug(f"Row {row}: Skipping WBE_D='{wbe_val}', WBE_C='{wbe_backup_val}', Offer={offer_val} (invalid format)")
            
            logger.info(f"Processed {processed_rows} rows, found {valid_offer_rows} rows with valid offers")
            logger.info(f"Successfully extracted {len(wbe_offers)} unique WBE codes with summed offers from VA21")
            
            # Log summary of extracted WBE offers
            total_extracted_offer = sum(wbe_offers.values())
            logger.info(f"Total offer value extracted: €{total_extracted_offer:,.2f}")
            
            # Log first few WBE mappings for debugging
            wbe_items = list(wbe_offers.items())
            for i, (wbe, offer) in enumerate(wbe_items[:5]):
                logger.info(f"  WBE '{wbe}' -> €{offer:,.2f}")
            if len(wbe_offers) > 5:
                logger.info(f"  ... and {len(wbe_offers) - 5} more WBE codes")
            
            # Check for any WBEs that have multiple entries (were summed)
            wbe_counts = {}
            for row in range(VA21Rows.DATA_START_ROW, va21_ws.max_row + 1):
                wbe_cell = va21_ws.cell(row=row, column=VA21Columns.WBE)
                wbe_backup_cell = va21_ws.cell(row=row, column=VA21Columns.WBE_BACKUP)
                offer_cell = va21_ws.cell(row=row, column=VA21Columns.OFFER_TOTAL)
                
                wbe_val = wbe_cell.value
                wbe_backup_val = wbe_backup_cell.value
                offer_val = offer_cell.value
                
                final_wbe = None
                if wbe_val and str(wbe_val).strip() and str(wbe_val).strip() != 'None':
                    final_wbe = str(wbe_val).strip()
                elif wbe_backup_val and str(wbe_backup_val).strip() and str(wbe_backup_val).strip() != 'None':
                    final_wbe = self._convert_wbe_us_to_it(str(wbe_backup_val).strip())
                
                if final_wbe and offer_val is not None and isinstance(offer_val, (int, float)):
                    wbe_counts[final_wbe] = wbe_counts.get(final_wbe, 0) + 1
            
            # Log WBEs that appeared multiple times
            duplicated_wbes = {wbe: count for wbe, count in wbe_counts.items() if count > 1}
            if duplicated_wbes:
                logger.info(f"Found {len(duplicated_wbes)} WBE codes with multiple entries (values were summed):")
                for wbe, count in list(duplicated_wbes.items())[:5]:
                    logger.info(f"  WBE '{wbe}': {count} entries, total €{wbe_offers[wbe]:,.2f}")
                if len(duplicated_wbes) > 5:
                    logger.info(f"  ... and {len(duplicated_wbes) - 5} more duplicated WBEs")
            else:
                logger.info("All WBE codes had unique entries (no duplication/summing needed)")
            
            return wbe_offers
            
        except Exception as e:
            logger.error(f"Error extracting VA21 data from sheet '{latest_sheet}': {e}")
            return {}

    def integrate_va21_offers_into_categories(self, product_groups: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Integrate VA21 offer prices into the parsed categories.
        Merge VA21 data into existing WBE categories and only create new categories for truly unmapped WBEs.
        
        Args:
            product_groups: List of product groups with categories
            
        Returns:
            Updated product groups with offer prices and merged VA21 data
        """
        # Extract VA21 offer data
        va21_offers = self.extract_va21_offer_data()
        
        if not va21_offers:
            logger.warning("No VA21 offer data available, categories will not have offer prices")
            return product_groups
        
        logger.info(f"Starting integration of VA21 offers into {sum(len(g.get('categories', [])) for g in product_groups)} categories")
        
        matched_offers = 0
        total_matched_value = 0
        existing_wbes = set()
        merged_wbes = set()  # Track WBEs that have been merged with VA21 data
        
        # First pass: integrate offer prices into existing categories and merge VA21 data
        for group in product_groups:
            for category in group[JsonFields.CATEGORIES]:
                category_wbe = category.get(JsonFields.WBE, "")
                
                if category_wbe:
                    existing_wbes.add(category_wbe)
                    
                    # Look up offer price directly in VA21 data
                    offer_price = va21_offers.get(category_wbe, 0.0)
                    
                    # Add offer price to category
                    category[JsonFields.OFFER_PRICE] = offer_price
                    
                    if offer_price > 0:
                        matched_offers += 1
                        total_matched_value += offer_price
                        merged_wbes.add(category_wbe)
                        logger.info(f"✓ Matched category WBE '{category_wbe}' -> Offer: €{offer_price:,.2f}")
                        
                        # Merge VA21 data into existing category items if available
                        self._merge_va21_data_into_category(category, category_wbe)
                    else:
                        logger.warning(f"✗ No offer price found for category WBE '{category_wbe}'")
                else:
                    # No WBE code, set offer price to 0
                    category[JsonFields.OFFER_PRICE] = 0.0
                    logger.debug("Category has no WBE code, setting offer price to 0")
        
        # Second pass: identify truly unmapped WBEs from VA21 (not merged in first pass)
        unmapped_wbes = {wbe: offer for wbe, offer in va21_offers.items() 
                        if wbe not in merged_wbes and wbe not in existing_wbes}
        
        if unmapped_wbes:
            logger.info(f"Found {len(unmapped_wbes)} truly unmapped WBE codes in VA21, creating new categories")
            
            # Get VA21 worksheet and headers for data extraction
            latest_sheet = self._find_latest_va21_sheet()
            if latest_sheet:
                va21_ws = self.workbook[latest_sheet]
                headers = self.extract_va21_headers(va21_ws)
                
                # Create a new group for VA21-only categories
                va21_group = {
                    JsonFields.GROUP_ID: "TXT-VA21",
                    JsonFields.GROUP_NAME: "Categories from VA21 (not in NEW_OFFER1)",
                    JsonFields.QUANTITY: 1,
                    JsonFields.CATEGORIES: []
                }
                
                # Create categories for each truly unmapped WBE
                for wbe_code, offer_price in unmapped_wbes.items():
                    try:
                        new_category = self.create_category_from_va21_wbe(wbe_code, offer_price, va21_ws, headers)
                        va21_group[JsonFields.CATEGORIES].append(new_category)
                        matched_offers += 1
                        total_matched_value += offer_price
                        logger.info(f"✓ Created new category for unmapped WBE '{wbe_code}' -> Offer: €{offer_price:,.2f}")
                    except Exception as e:
                        logger.error(f"Failed to create category for WBE '{wbe_code}': {e}")
                
                # Add the VA21 group if it has categories
                if va21_group[JsonFields.CATEGORIES]:
                    product_groups.append(va21_group)
                    logger.info(f"Added new group 'TXT-VA21' with {len(va21_group[JsonFields.CATEGORIES])} categories from VA21")
        else:
            logger.info("All VA21 WBE codes were successfully merged into existing categories")
        
        # Verify total offer prices match
        final_total_offers = sum(va21_offers.values())
        logger.info(f"Integration completed: {matched_offers} total categories with total value €{total_matched_value:,.2f}")
        logger.info(f"VA21 total: €{final_total_offers:,.2f}, Integrated total: €{total_matched_value:,.2f}")
        logger.info(f"Merged into existing: {len(merged_wbes)}, New categories: {len(unmapped_wbes)}")
        
        if abs(final_total_offers - total_matched_value) > 0.01:  # Allow for small rounding differences
            logger.warning(f"Total mismatch: VA21 total €{final_total_offers:,.2f} != Integrated total €{total_matched_value:,.2f}")
        else:
            logger.info("✓ Total offer prices match between VA21 and integrated categories")
        
        return product_groups

    def _merge_va21_data_into_category(self, category: Dict[str, Any], wbe_code: str):
        """
        Merge additional VA21 data into an existing category from NEW_OFFER1.
        
        Args:
            category: Existing category dictionary to enhance
            wbe_code: WBE code to look up in VA21
        """
        try:
            latest_sheet = self._find_latest_va21_sheet()
            if not latest_sheet:
                return
            
            va21_ws = self.workbook[latest_sheet]
            headers = self.extract_va21_headers(va21_ws)
            
            # Find VA21 rows for this WBE and extract additional data
            va21_items = []
            for row in range(VA21Rows.DATA_START_ROW, va21_ws.max_row + 1):
                # Check both Column D and Column C for WBE
                wbe_cell_d = va21_ws.cell(row=row, column=VA21Columns.WBE)
                wbe_cell_c = va21_ws.cell(row=row, column=VA21Columns.WBE_BACKUP)
                
                wbe_val_d = wbe_cell_d.value
                wbe_val_c = wbe_cell_c.value
                
                # Determine WBE for this row
                row_wbe = None
                if wbe_val_d and str(wbe_val_d).strip() and str(wbe_val_d).strip() != 'None':
                    row_wbe = str(wbe_val_d).strip()
                elif wbe_val_c and str(wbe_val_c).strip() and str(wbe_val_c).strip() != 'None':
                    row_wbe = self._convert_wbe_us_to_it(str(wbe_val_c).strip())
                
                if row_wbe == wbe_code:
                    # Extract item data from this VA21 row
                    item_data = self.extract_va21_row_data(va21_ws, row, headers)
                    if item_data.get(JsonFields.DESCRIPTION):  # Only add if has description
                        # Mark this as VA21 source data
                        item_data['va21_source'] = True
                        item_data[JsonFields.POSITION] = f"VA21-{row}"
                        va21_items.append(item_data)
            
            # Add VA21 items to the existing category
            if va21_items:
                category[JsonFields.ITEMS].extend(va21_items)
                logger.debug(f"Merged {len(va21_items)} VA21 items into existing category '{wbe_code}'")
                
                # Recalculate category totals to include VA21 data
                total_listino = sum(item.get(JsonFields.PRICELIST_TOTAL, 0) for item in category[JsonFields.ITEMS])
                total_cost = sum(item.get(JsonFields.TOTAL_COST, 0) for item in category[JsonFields.ITEMS])
                
                # Update category totals (but preserve original subtotals from NEW_OFFER1)
                category['total_listino_with_va21'] = total_listino
                category['total_cost_with_va21'] = total_cost
                
        except Exception as e:
            logger.warning(f"Failed to merge VA21 data for WBE '{wbe_code}': {e}")

    def extract_va21_headers(self, va21_ws) -> Dict[int, str]:
        """
        Extract column headers from VA21 sheet row 18.
        
        Args:
            va21_ws: VA21 worksheet
            
        Returns:
            Dictionary mapping column index to header name
        """
        headers = {}
        header_row = VA21Rows.HEADER_ROW
        
        # Extract headers starting from column A
        for col in range(1, va21_ws.max_column + 1):
            cell = va21_ws.cell(row=header_row, column=col)
            if cell.value and isinstance(cell.value, str):
                header_name = cell.value.strip()
                if header_name:
                    headers[col] = header_name
        
        logger.info(f"Extracted {len(headers)} headers from VA21 row {header_row}")
        return headers

    def extract_va21_row_data(self, va21_ws, row: int, headers: Dict[int, str]) -> Dict[str, Any]:
        """
        Extract data from a VA21 row and map it to NEW_OFFER1 fields.
        
        Args:
            va21_ws: VA21 worksheet
            row: Row number to extract
            headers: Column headers mapping
            
        Returns:
            Dictionary with mapped field data
        """
        row_data = {}
        
        # Extract raw data from VA21 row
        for col, header_name in headers.items():
            cell = va21_ws.cell(row=row, column=col)
            
            # Map VA21 field to NEW_OFFER1 field if mapping exists
            if header_name in VA21FieldMapping.MAPPINGS:
                new_offer_field = VA21FieldMapping.MAPPINGS[header_name]
                
                # Convert cell value based on field type
                if cell.value is not None:
                    if new_offer_field in [JsonFields.QTY, JsonFields.PRICELIST_TOTAL, JsonFields.PRICELIST_UNIT_PRICE]:
                        # Numeric fields
                        row_data[new_offer_field] = self._safe_float(cell.value)
                    else:
                        # Text fields
                        row_data[new_offer_field] = str(cell.value).strip()
        
        # Set default values for missing fields
        row_data.setdefault(JsonFields.POSITION, str(row))
        row_data.setdefault(JsonFields.QTY, 1.0)
        row_data.setdefault(JsonFields.PRICELIST_TOTAL, 0.0)
        row_data.setdefault(JsonFields.PRICELIST_UNIT_PRICE, 0.0)
        row_data.setdefault(JsonFields.UNIT_COST, 0.0)
        row_data.setdefault(JsonFields.TOTAL_COST, 0.0)
        
        # Set all other fields to default values
        for field_name in [JsonFields.COD_LISTINO, JsonFields.INTERNAL_CODE, JsonFields.PRIORITY_ORDER,
                          JsonFields.PRIORITY, JsonFields.LINE_NUMBER, JsonFields.WBS, JsonFields.TOTAL]:
            row_data.setdefault(field_name, "")
        
        # Set all numeric fields to 0
        numeric_fields = [
            JsonFields.MAT, JsonFields.UTM_ROBOT, JsonFields.UTM_ROBOT_H, JsonFields.UTM_LGV, JsonFields.UTM_LGV_H,
            JsonFields.UTM_INTRA, JsonFields.UTM_INTRA_H, JsonFields.UTM_LAYOUT, JsonFields.UTM_LAYOUT_H,
            JsonFields.UTE, JsonFields.UTE_H, JsonFields.BA, JsonFields.BA_H, JsonFields.SW_PC, JsonFields.SW_PC_H,
            JsonFields.SW_PLC, JsonFields.SW_PLC_H, JsonFields.SW_LGV, JsonFields.SW_LGV_H,
            JsonFields.MTG_MEC, JsonFields.MTG_MEC_H, JsonFields.MTG_MEC_INTRA, JsonFields.MTG_MEC_INTRA_H,
            JsonFields.CAB_ELE, JsonFields.CAB_ELE_H, JsonFields.CAB_ELE_INTRA, JsonFields.CAB_ELE_INTRA_H,
            JsonFields.COLL_BA, JsonFields.COLL_BA_H, JsonFields.COLL_PC, JsonFields.COLL_PC_H,
            JsonFields.COLL_PLC, JsonFields.COLL_PLC_H, JsonFields.COLL_LGV, JsonFields.COLL_LGV_H,
            JsonFields.PM_COST, JsonFields.PM_H, JsonFields.SPESE_PM, JsonFields.DOCUMENT, JsonFields.DOCUMENT_H,
            JsonFields.IMBALLO, JsonFields.STOCCAGGIO, JsonFields.TRASPORTO, JsonFields.SITE, JsonFields.SITE_H,
            JsonFields.INSTALL, JsonFields.INSTALL_H, JsonFields.AV_PC, JsonFields.AV_PC_H,
            JsonFields.AV_PLC, JsonFields.AV_PLC_H, JsonFields.AV_LGV, JsonFields.AV_LGV_H,
            JsonFields.SPESE_FIELD, JsonFields.SPESE_VARIE, JsonFields.AFTER_SALES,
            JsonFields.PROVVIGIONI_ITALIA, JsonFields.PROVVIGIONI_ESTERO, JsonFields.TESORETTO,
            JsonFields.MONTAGGIO_BEMA_MBE_US
        ]
        
        for field in numeric_fields:
            row_data.setdefault(field, 0.0)
        
        return row_data

    def create_category_from_va21_wbe(self, wbe_code: str, offer_price: float, va21_ws, headers: Dict[int, str]) -> Dict[str, Any]:
        """
        Create a category from VA21 data for a WBE that doesn't exist in NEW_OFFER1.
        
        Args:
            wbe_code: WBE code to create category for
            offer_price: Total offer price for this WBE
            va21_ws: VA21 worksheet
            headers: Column headers mapping
            
        Returns:
            Dictionary representing the new category
        """
        # Find all rows in VA21 that belong to this WBE
        wbe_rows = []
        items = []
        
        for row in range(VA21Rows.DATA_START_ROW, va21_ws.max_row + 1):
            # Check both Column D and Column C for WBE
            wbe_cell_d = va21_ws.cell(row=row, column=VA21Columns.WBE)
            wbe_cell_c = va21_ws.cell(row=row, column=VA21Columns.WBE_BACKUP)
            
            wbe_val_d = wbe_cell_d.value
            wbe_val_c = wbe_cell_c.value
            
            # Determine WBE for this row
            row_wbe = None
            if wbe_val_d and str(wbe_val_d).strip() and str(wbe_val_d).strip() != 'None':
                row_wbe = str(wbe_val_d).strip()
            elif wbe_val_c and str(wbe_val_c).strip() and str(wbe_val_c).strip() != 'None':
                row_wbe = self._convert_wbe_us_to_it(str(wbe_val_c).strip())
            
            if row_wbe == wbe_code:
                wbe_rows.append(row)
                
                # Extract item data from this row
                item_data = self.extract_va21_row_data(va21_ws, row, headers)
                if item_data.get(JsonFields.DESCRIPTION):  # Only add if has description
                    items.append(item_data)
        
        # If no items found, create a dummy item
        if not items:
            items = [{
                JsonFields.POSITION: "1",
                JsonFields.CODE: wbe_code,
                JsonFields.COD_LISTINO: "",
                JsonFields.DESCRIPTION: f"Item from VA21 for WBE {wbe_code}",
                JsonFields.QTY: 1.0,
                JsonFields.PRICELIST_UNIT_PRICE: offer_price,
                JsonFields.PRICELIST_TOTAL: offer_price,
                JsonFields.UNIT_COST: 0.0,
                JsonFields.TOTAL_COST: 0.0,
                **{field: 0.0 for field in [
                    JsonFields.MAT, JsonFields.UTM_ROBOT, JsonFields.UTM_ROBOT_H, JsonFields.UTM_LGV, JsonFields.UTM_LGV_H,
                    JsonFields.UTM_INTRA, JsonFields.UTM_INTRA_H, JsonFields.UTM_LAYOUT, JsonFields.UTM_LAYOUT_H,
                    JsonFields.UTE, JsonFields.UTE_H, JsonFields.BA, JsonFields.BA_H, JsonFields.SW_PC, JsonFields.SW_PC_H,
                    JsonFields.SW_PLC, JsonFields.SW_PLC_H, JsonFields.SW_LGV, JsonFields.SW_LGV_H,
                    JsonFields.MTG_MEC, JsonFields.MTG_MEC_H, JsonFields.MTG_MEC_INTRA, JsonFields.MTG_MEC_INTRA_H,
                    JsonFields.CAB_ELE, JsonFields.CAB_ELE_H, JsonFields.CAB_ELE_INTRA, JsonFields.CAB_ELE_INTRA_H,
                    JsonFields.COLL_BA, JsonFields.COLL_BA_H, JsonFields.COLL_PC, JsonFields.COLL_PC_H,
                    JsonFields.COLL_PLC, JsonFields.COLL_PLC_H, JsonFields.COLL_LGV, JsonFields.COLL_LGV_H,
                    JsonFields.PM_COST, JsonFields.PM_H, JsonFields.SPESE_PM, JsonFields.DOCUMENT, JsonFields.DOCUMENT_H,
                    JsonFields.IMBALLO, JsonFields.STOCCAGGIO, JsonFields.TRASPORTO, JsonFields.SITE, JsonFields.SITE_H,
                    JsonFields.INSTALL, JsonFields.INSTALL_H, JsonFields.AV_PC, JsonFields.AV_PC_H,
                    JsonFields.AV_PLC, JsonFields.AV_PLC_H, JsonFields.AV_LGV, JsonFields.AV_LGV_H,
                    JsonFields.SPESE_FIELD, JsonFields.SPESE_VARIE, JsonFields.AFTER_SALES,
                    JsonFields.PROVVIGIONI_ITALIA, JsonFields.PROVVIGIONI_ESTERO, JsonFields.TESORETTO,
                    JsonFields.MONTAGGIO_BEMA_MBE_US
                ]}
            }]
        
        # Calculate category totals
        total_listino = sum(item.get(JsonFields.PRICELIST_TOTAL, 0) for item in items)
        total_cost = sum(item.get(JsonFields.TOTAL_COST, 0) for item in items)
        
        # Create category
        category = {
            JsonFields.CATEGORY_ID: wbe_code.split('-')[-2] if '-' in wbe_code else wbe_code[:4],  # Extract category from WBE
            JsonFields.CATEGORY_CODE: wbe_code,
            JsonFields.CATEGORY_NAME: f"Category from VA21 - {wbe_code}",
            JsonFields.WBE: wbe_code,
            JsonFields.PRICELIST_SUBTOTAL: total_listino,
            JsonFields.COST_SUBTOTAL: total_cost,
            JsonFields.TOTAL_COST: total_cost,
            JsonFields.OFFER_PRICE: offer_price,
            JsonFields.ITEMS: items
        }
        
        logger.info(f"Created new category for VA21 WBE '{wbe_code}' with {len(items)} items and offer €{offer_price:,.2f}")
        return category

def parse_analisi_profittabilita_to_json(file_path: str, output_path: Optional[str] = None) -> Dict[str, Any]:
    """
    Main function to parse Analisi Profittabilita Excel file and optionally save to JSON
    
    Args:
        file_path: Path to the Excel file
        output_path: Optional path to save JSON output
        
    Returns:
        Dictionary containing the parsed data
    """
    parser = AnalisiProfittabilitaParser(file_path)
    result = parser.parse()
    
    if output_path:
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, indent=2, ensure_ascii=False)
        logger.info(LogMessages.JSON_SAVED.format(output_path))
    
    return result

def parse_analisi_profittabilita_to_model(file_path: str, output_path: Optional[str] = None) -> IndustrialQuotation:
    """
    Main function to parse Analisi Profittabilita Excel file to IndustrialQuotation model with validation
    
    Args:
        file_path: Path to the Excel file
        output_path: Optional path to save JSON output using model serialization
        
    Returns:
        IndustrialQuotation: Validated quotation model instance
    """
    parser = AnalisiProfittabilitaParser(file_path)
    quotation = parser.parse_to_model()
    
    if output_path:
        quotation.save_json(output_path)
    
    return quotation


if __name__ == "__main__":
    # Example usage
    input_file = r"input\TEST_Analisi profitabilita'_Tabella riassuntiva SAP_PEPSICO INC - Copia.xlsm"
    output_file = "output/analisi_profittabilita_with_va21_offers.json"
    
    try:
        result = parse_analisi_profittabilita_to_json(input_file, output_file)
        print("=" * 80)
        print("ANALISI PROFITTABILITÀ - RESULTS WITH VA21 OFFER INTEGRATION")
        print("=" * 80)
        print(f"Project ID: {result['project']['id']}")
        print(f"Listino: {result['project']['listino']}")
        print(f"Number of product groups: {len(result['product_groups'])}")
        print()
        
        # Financial Summary
        totals = result['totals']
        print("FINANCIAL SUMMARY:")
        print("-" * 40)
        print(f"Total Listino:        €{totals['total_listino']:>15,.2f}")
        print(f"Total Cost:           €{totals['total_costo']:>15,.2f}")
        print(f"Total Offer:          €{totals['total_offer']:>15,.2f}")
        print()
        print(f"Listino Margin:       €{totals['margin']:>15,.2f} ({totals['margin_percentage']:>6.2f}%)")
        print(f"Offer Margin:         €{totals['offer_margin']:>15,.2f} ({totals['offer_margin_percentage']:>6.2f}%)")
        print()
        
        # Show categories with offer prices and margins
        print("CATEGORIES WITH VA21 OFFER PRICES:")
        print("-" * 80)
        categories_with_offers = []
        
        for group in result['product_groups']:
            for category in group['categories']:
                offer_price = category.get('offer_price', 0)
                if offer_price > 0:
                    cost = category.get('cost_subtotal', 0)
                    # Calculate margin using corrected formula: margin% = 1 - (cost / offer)
                    margin_amount = offer_price - cost
                    margin_pct = (1 - (cost / offer_price)) * 100 if offer_price > 0 else 0
                    
                    categories_with_offers.append({
                        'wbe': category.get('wbe', 'N/A'),
                        'name': category.get('category_name', 'N/A'),
                        'offer_price': offer_price,
                        'cost': cost,
                        'margin_amount': margin_amount,
                        'margin_pct': margin_pct
                    })
        
        if categories_with_offers:
            for i, cat in enumerate(categories_with_offers[:5], 1):  # Show first 5
                print(f"{i:2d}. {cat['name']}")
                print(f"    WBE:          {cat['wbe']}")
                print(f"    Offer Price:  €{cat['offer_price']:>12,.2f}")
                print(f"    Cost:         €{cat['cost']:>12,.2f}")
                print(f"    Margin:       €{cat['margin_amount']:>12,.2f} ({cat['margin_pct']:>6.2f}%)")
                print()
            
            if len(categories_with_offers) > 5:
                print(f"... and {len(categories_with_offers) - 5} more categories with offer prices")
        else:
            print("No categories with offer prices found")
        
        print("=" * 80)
        print(f"Results saved to: {output_file}")
        
    except Exception as e:
        logger.error(f"Error parsing Excel file: {e}")
        raise 