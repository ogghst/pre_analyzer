#!/usr/bin/env python3
"""
Direct Analisi Profittabilita Parser - Excel to IndustrialQuotation Model
Converts Excel files with Analisi Profittabilita format directly to IndustrialQuotation objects
"""

import logging
import decimal
from typing import List, Optional, Any, Dict
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from openpyxl import load_workbook

# Import unified models
import sys
import os
sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from models.quotation_models import (
    IndustrialQuotation, ProjectInfo, ProjectParameters, SalesInfo,
    ProductGroup, QuotationCategory, QuotationItem, QuotationTotals,
    CurrencyType, CategoryType, ParserType
)

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

from utils.format import safe_format_number

# =============================================================================
# CONSTANTS
# =============================================================================

# Excel Column Indices (1-based) - Complete mapping for all 81 columns
class ExcelColumns:
    # Basic project columns (1-21)
    COD = 1                          # COD
    PRIORITY_ORDER = 2               # Priorità per ordinamento
    PRIORITY = 3                     # Priorità
    LINE_NUMBER = 4                  # NUMERA LE RIGHE
    WBS = 5                          # WBS
    WBE = 6                          # WBE
    POSITION = 7                     # POS.
    CODICE = 8                       # CODICE
    COD_LISTINO = 9                  # COD.LISTINO
    DENOMINAZIONE = 10               # DENOMINAZIONE
    QTA = 11                         # QTA'
    SUB_TOT_LISTINO = 12             # SUB TOT. LISTINO
    LIST_UNIT = 13                   # LIST.UNIT.
    LISTINO_TOTALE = 14              # LISTINO TOTALE
    SUBTOT_COSTO = 15                # SUBTOT COSTO
    COSTO_UNITARIO = 16              # COSTO UNITARIO
    COSTO_TOTALE = 17                # COSTO TOTALE
    # Column 18 is empty
    COD_2 = 19                       # COD (second occurrence)
    # Column 20 is empty
    TOTALE = 21                      # TOTALE
    
    # Material and UTM columns (22-30)
    MAT = 22                         # MAT
    UTM_ROBOT = 23                   # UTM-ROBOT
    UTM_ROBOT_H = 24                 # UTM-ROBOT-H
    UTM_LGV = 25                     # UTM-LGV
    UTM_LGV_H = 26                   # UTM-LGV-H
    UTM_INTRA = 27                   # UTM-INTRA
    UTM_INTRA_H = 28                 # UTM-INTRA-H
    UTM_LAYOUT = 29                  # UTM-LAYOUT
    UTM_LAYOUT_H = 30                # UTM-LAYOUT-H
    
    # Engineering columns (31-40)
    UTE = 31                         # UTE
    UTE_H = 32                       # UTE-H
    BA = 33                          # BA
    BA_H = 34                        # BA-H
    SW_PC = 35                       # SW-PC
    SW_PC_H = 36                     # SW-PC-H
    SW_PLC = 37                      # SW-PLC
    SW_PLC_H = 38                    # SW-PLC-H
    SW_LGV = 39                      # SW-LGV
    SW_LGV_H = 40                    # SW-LGV-H
    
    # Manufacturing columns (41-50)
    MTG_MEC = 41                     # MTG-MEC
    MTG_MEC_H = 42                   # MTG-MEC-H
    MTG_MEC_INTRA = 43               # MTG-MEC-INTRA
    MTG_MEC_INTRA_H = 44             # MTG-MEC-INTRA-H
    CAB_ELE = 45                     # CAB-ELE
    CAB_ELE_H = 46                   # CAB-ELE-H
    CAB_ELE_INTRA = 47               # CAB-ELE-INTRA
    CAB_ELE_INTRA_H = 48             # CAB-ELE-INTRA-H
    COLL_BA = 49                     # COLL-BA
    COLL_BA_H = 50                   # COLL-BA-H
    
    # Testing columns (51-60)
    COLL_PC = 51                     # COLL-PC
    COLL_PC_H = 52                   # COLL-PC-H
    COLL_PLC = 53                    # COLL-PLC
    COLL_PLC_H = 54                  # COLL-PLC-H
    COLL_LGV = 55                    # COLL-LGV
    COLL_LGV_H = 56                  # COLL-LGV-H
    PM_COST = 57                     # PM
    PM_H = 58                        # PM-H
    SPESE_PM = 59                    # SPESE-PM
    DOCUMENT = 60                    # DOCUMENT
    
    # Logistics and field columns (61-70)
    DOCUMENT_H = 61                  # DOCUMENT-H
    IMBALLO = 62                     # IMBALLO
    STOCCAGGIO = 63                  # STOCCAGGIO
    TRASPORTO = 64                   # TRASPORTO
    SITE = 65                        # SITE
    SITE_H = 66                      # SITE-H
    INSTALL = 67                     # INSTALL
    INSTALL_H = 68                   # INSTALL-H
    AV_PC = 69                       # AV-PC
    AV_PC_H = 70                     # AV-PC-H
    
    # Additional field columns (71-81)
    AV_PLC = 71                      # AV-PLC
    AV_PLC_H = 72                    # AV-PLC-H
    AV_LGV = 73                      # AV-LGV
    AV_LGV_H = 74                    # AV-LGV-H
    SPESE_FIELD = 75                 # SPESE FIELD
    SPESE_VARIE = 76                 # SPESE VARIE
    AFTER_SALES = 77                 # AFTER SALES
    PROVVIGIONI_ITALIA = 78          # PROVVIGIONI ITALIA
    PROVVIGIONI_ESTERO = 79          # PROVVIGIONI ESTERO
    TESORETTO = 80                   # TESORETTO
    MONTAGGIO_BEMA_MBE_US = 81       # MONTAGGIO BEMA MBE-US

# VA21 Sheet Constants
class VA21Columns:
    WBE_BACKUP = 3                   # Column C - WBE codes (backup when Column D is None)
    WBE = 4                          # Column D - WBE codes (primary)
    COD = 2                          
    DESCRIPTION = 11       
    QUANTITY = 13       
    LISTINO_SUBTOTAL = 22
    DISCOUNT = 24
    OFFER_TOTAL = 25                 # Column Y - Offer totals
    COST_SUBTOTAL = 27
    MARGIN_PERCENTAGE = 28

class VA21Rows:
    DATA_START_ROW = 19              # Data starts from row 18
    HEADER_ROW = 18                  # Headers are in row 18

# Excel Row Constants
class ExcelRows:
    HEADER_ROW = 3
    DATA_START_ROW = 4
    PROJECT_INFO_START = 1
    PROJECT_INFO_END = 6

# Project Information Cell Positions (row, column)
class ProjectInfoCells:
    PROJECT_ID = (1, 1)        # Row 1, Column A - Project identifier
    LISTINO = (2, 1)           # Row 2, Column A - Price list info

# Identification Patterns
class IdentificationPatterns:
    GROUP_PREFIX = 'TXT'
    CATEGORY_CODE_LENGTH = 4
    HEADER_CODE = 'COD'
    VA21_SHEET_PREFIX = 'VA21'       # Prefix for VA21 sheets
    WBE_IT_SUFFIX = '-IT'            # Italian WBE suffix in NEW_OFFER1
    WBE_US_SUFFIX = '-US'            # US WBE suffix in VA21 sheets

# =============================================================================
# DIRECT ANALISI PROFITTABILITA PARSER
# =============================================================================

class DirectAnalisiProfittabilitaParser:
    """Direct parser for converting Analisi Profittabilita Excel files to IndustrialQuotation objects"""
    
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.workbook = None
        self.ws = None
        self.va21_offers = {}  # Cache for VA21 offer data
        
    def load_workbook(self) -> None:
        """Load the Excel workbook"""
        try:
            logger.info(f"Loading workbook from: {self.file_path}")
            self.workbook = load_workbook(str(self.file_path), data_only=True)
            logger.debug("Workbook loaded successfully")
            
            if 'NEW_OFFER1' not in self.workbook.sheetnames:
                error_msg = f"Required worksheet 'NEW_OFFER1' not found in {self.file_path}. Available sheets: {self.workbook.sheetnames}"
                logger.error(error_msg)
                raise ValueError(error_msg)
            
            self.ws = self.workbook['NEW_OFFER1']
            logger.info(f"Loaded workbook with {self.ws.max_row} rows and {self.ws.max_column} columns")
            
        except FileNotFoundError as e:
            error_msg = f"File not found: {self.file_path}"
            logger.error(error_msg)
            raise FileNotFoundError(error_msg) from e
        except PermissionError as e:
            error_msg = f"Permission denied accessing file: {self.file_path}"
            logger.error(error_msg)
            raise PermissionError(error_msg) from e
        except ValueError as e:
            # Re-raise ValueError with original exception
            raise e
        except Exception as e:
            error_msg = f"Unable to load workbook {self.file_path}: {str(e)}"
            logger.error(error_msg)
            raise Exception(error_msg) from e
    
    def extract_project_info(self) -> ProjectInfo:
        """Extract project information and create ProjectInfo object"""
        try:
            logger.debug("Extracting project information")
            
            if not self.ws:
                error_msg = "Worksheet not loaded. Call load_workbook() first."
                logger.error(error_msg)
                raise RuntimeError(error_msg)
            
            # Extract basic project info with error handling
            try:
                project_id_cell = self.ws.cell(row=ProjectInfoCells.PROJECT_ID[0], column=ProjectInfoCells.PROJECT_ID[1])
                project_id = str(project_id_cell.value) if project_id_cell.value else ""
            except Exception as e:
                logger.warning(f"Error extracting project ID: {e}")
                project_id = ""
            
            try:
                listino_cell = self.ws.cell(row=ProjectInfoCells.LISTINO[0], column=ProjectInfoCells.LISTINO[1])
                listino = str(listino_cell.value) if listino_cell.value else None
            except Exception as e:
                logger.warning(f"Error extracting listino: {e}")
                listino = None
            
            # Create parameters object with defaults for AP files
            try:
                parameters = ProjectParameters(
                    doc_percentage=0.00632,  # Default value
                    pm_percentage=0.02061,   # Default value
                    financial_costs=0.0,
                    currency="EUR",          # Default currency
                    exchange_rate=1.0,
                    waste_disposal=0.005,
                    warranty_percentage=0.03,
                    is_24h_service=False
                )
            except Exception as e:
                error_msg = f"Error creating ProjectParameters: {e}"
                logger.error(error_msg)
                raise ValueError(error_msg) from e
            
            # Create sales info object (defaults for AP files)
            try:
                sales_info = SalesInfo(
                    area_manager=None,
                    agent=None,
                    commission_percentage=0.0,
                    author=None
                )
            except Exception as e:
                error_msg = f"Error creating SalesInfo: {e}"
                logger.error(error_msg)
                raise ValueError(error_msg) from e
            
            try:
                project_info = ProjectInfo(
                    id=project_id,
                    customer="",  # Not typically available in AP files
                    listino=listino,
                    parameters=parameters,
                    sales_info=sales_info
                )
                logger.debug(f"Successfully extracted project info: ID={project_id}")
                return project_info
            except Exception as e:
                error_msg = f"Error creating ProjectInfo: {e}"
                logger.error(error_msg)
                raise ValueError(error_msg) from e
                
        except RuntimeError as e:
            # Re-raise runtime errors
            raise e
        except ValueError as e:
            # Re-raise value errors
            raise e
        except Exception as e:
            error_msg = f"Unexpected error extracting project info: {e}"
            logger.error(error_msg)
            raise Exception(error_msg) from e
    
    def extract_va21_offer_data(self) -> Dict[str, float]:
        """Extract offer prices from VA21 sheet if available"""
        try:
            logger.debug("Starting VA21 offer data extraction")
            
            if not self.workbook:
                error_msg = "Workbook not loaded. Call load_workbook() first."
                logger.error(error_msg)
                raise RuntimeError(error_msg)
            
            try:
                va21_sheet_name = self._find_latest_va21_sheet()
                if not va21_sheet_name:
                    logger.debug("No VA21 sheet found")
                    return {}
            except Exception as e:
                error_msg = f"Error finding VA21 sheet: {e}"
                logger.error(error_msg)
                raise Exception(error_msg) from e
            
            try:
                va21_ws = self.workbook[va21_sheet_name]
                logger.debug(f"Processing VA21 sheet: {va21_sheet_name}")
                
                offers = {}
                for row in range(VA21Rows.DATA_START_ROW, va21_ws.max_row + 1):
                    try:
                        # Get WBE code (try column D first, then column C as backup)
                        wbe_primary = va21_ws.cell(row=row, column=VA21Columns.WBE).value
                        wbe_backup = va21_ws.cell(row=row, column=VA21Columns.WBE_BACKUP).value
                        wbe_code = wbe_primary if wbe_primary else wbe_backup
                        cod = va21_ws.cell(row=row, column=VA21Columns.COD).value
                        description = va21_ws.cell(row=row, column=VA21Columns.DESCRIPTION).value
                        quantity = va21_ws.cell(row=row, column=VA21Columns.QUANTITY).value
                        listino_subtotal = va21_ws.cell(row=row, column=VA21Columns.LISTINO_SUBTOTAL).value
                        discount = va21_ws.cell(row=row, column=VA21Columns.DISCOUNT).value
                        offer_total = va21_ws.cell(row=row, column=VA21Columns.OFFER_TOTAL).value
                        cost_subtotal = va21_ws.cell(row=row, column=VA21Columns.COST_SUBTOTAL).value
                        margin_percentage = va21_ws.cell(row=row, column=VA21Columns.MARGIN_PERCENTAGE).value
                        
                        if cod and wbe_code:
                            try:
                                wbe_str = str(wbe_code).strip()
                                # Convert US format to IT format if needed
                                if wbe_str.endswith(IdentificationPatterns.WBE_US_SUFFIX):
                                    wbe_str = self._convert_wbe_us_to_it(wbe_str)
                                
                                # Ensure offer_total is a valid number
                                safe_offer_total = self._safe_decimal(offer_total, 0.0)
                                safe_cost_subtotal = self._safe_decimal(cost_subtotal, 0.0)
                                safe_listino_subtotal = self._safe_decimal(listino_subtotal, 0.0)
                                safe_discount = self._safe_decimal(discount, 0.0)
                                safe_margin_percentage = self._safe_decimal(margin_percentage, 0.0)
                                
                                offers[wbe_str] = {
                                    VA21Columns.WBE: wbe_str,
                                    VA21Columns.COD: cod,
                                    VA21Columns.DESCRIPTION: description,
                                    VA21Columns.QUANTITY: quantity,
                                    VA21Columns.LISTINO_SUBTOTAL: safe_listino_subtotal,
                                    VA21Columns.DISCOUNT: safe_discount,
                                    VA21Columns.OFFER_TOTAL: safe_offer_total,
                                    VA21Columns.COST_SUBTOTAL: safe_cost_subtotal,
                                    VA21Columns.MARGIN_PERCENTAGE: safe_margin_percentage
                                }
                                logger.debug(f"Extracted VA21 offer: {wbe_str}")
                                
                            except Exception as e:
                                logger.warning(f"Error processing VA21 row {row} data: {e}")
                                continue
                                
                    except Exception as e:
                        logger.warning(f"Error reading VA21 row {row}: {e}")
                        continue
                
                logger.debug(f"Extracted {len(offers)} offer prices from VA21 sheet")
                return offers
                
            except KeyError as e:
                error_msg = f"VA21 sheet '{va21_sheet_name}' not found in workbook: {e}"
                logger.error(error_msg)
                raise KeyError(error_msg) from e
            except Exception as e:
                error_msg = f"Error processing VA21 sheet '{va21_sheet_name}': {e}"
                logger.error(error_msg)
                raise Exception(error_msg) from e
                
        except RuntimeError as e:
            # Re-raise runtime errors
            raise e
        except KeyError as e:
            # Re-raise key errors
            raise e
        except Exception as e:
            error_msg = f"Unexpected error extracting VA21 offer data: {e}"
            logger.error(error_msg)
            raise Exception(error_msg) from e
    
    def extract_product_groups(self) -> List[ProductGroup]:
        """Extract product groups, categories, and items directly as model objects"""
        try:
            logger.debug("Starting product groups extraction")
            
            if not self.ws:
                error_msg = "Worksheet not loaded. Call load_workbook() first."
                logger.error(error_msg)
                raise RuntimeError(error_msg)
            
            # First extract VA21 offer data
            try:
                self.va21_offers = self.extract_va21_offer_data()
            except Exception as e:
                logger.warning(f"Error extracting VA21 offers, continuing without them: {e}")
                self.va21_offers = {}
            
            product_groups: List[ProductGroup] = []
            current_group = None
            current_category = None
            
            # Start from data start row
            try:
                for row in range(ExcelRows.DATA_START_ROW, self.ws.max_row + 1):
                    try:
                        # Skip row if no priority value
                        priority_val = self._safe_cell_value(row, ExcelColumns.PRIORITY)
                        if not priority_val:
                            continue
                        
                        # Extract basic identification values
                        cod_val = self._safe_cell_value(row, ExcelColumns.COD)
                        codice_val = self._safe_cell_value(row, ExcelColumns.CODICE)
                        denominazione_val = self._safe_cell_value(row, ExcelColumns.DENOMINAZIONE)
                        qta_val = self._safe_cell_value(row, ExcelColumns.QTA)
                        wbe_val = self._safe_cell_value(row, ExcelColumns.WBE)

                        # Check if this is a group header (TXT in CODICE)
                        if codice_val and str(codice_val).startswith(IdentificationPatterns.GROUP_PREFIX):
                            try:
                                # Save previous group if exists
                                if current_group:
                                    product_groups.append(current_group)
                                
                                # Start new group
                                current_group = ProductGroup(
                                    group_id=str(codice_val),
                                    group_name=str(denominazione_val) if denominazione_val else "",
                                    quantity=self._safe_int(qta_val, 1),
                                    categories=[]
                                )
                                current_category = None
                                logger.debug(f"Found group: {codice_val}")
                            except Exception as e:
                                logger.error(f"Error creating product group for row {row}: {e}")
                                continue
                                
                        # Check if this is a category (4-char code in COD column)
                        elif cod_val and len(str(cod_val).strip()) == IdentificationPatterns.CATEGORY_CODE_LENGTH and current_group:
                            try:
                                # Get offer price from VA21 if available
                                wbe_code = str(wbe_val) if wbe_val else ""
                                offer_price = self.va21_offers.get(wbe_code, {}).get(VA21Columns.OFFER_TOTAL, 0.0)
                                
                                # Calculate cost value safely
                                cost_value = float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.SUBTOT_COSTO)))
                                
                                # Calculate margin safely, handling None offer_price
                                margin_amount = (offer_price - cost_value) if offer_price is not None else 0.0
                                
                                # Calculate margin percentage safely
                                margin_percentage = (
                                    ((offer_price - cost_value) / cost_value * 100)
                                    if offer_price is not None and cost_value != 0
                                    else 0.0
                                )
                                
                                current_category = QuotationCategory(
                                    category_id=str(cod_val),
                                    category_name=str(denominazione_val) if denominazione_val else "",
                                    wbe=wbe_code,
                                    items=[],
                                    pricelist_subtotal=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.SUB_TOT_LISTINO))),
                                    cost_subtotal=cost_value,
                                    total_cost=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.COSTO_TOTALE))),
                                    offer_price=offer_price,
                                    margin_amount=margin_amount,
                                    margin_percentage=margin_percentage
                                )
                                current_group.categories.append(current_category)
                                logger.debug(f"Found category: {current_category.category_id} - \
                                    list {safe_format_number(current_category.pricelist_subtotal,0)} - \
                                    cost {safe_format_number(current_category.cost_subtotal,0)} - \
                                    offer {safe_format_number(current_category.offer_price,0)} - \
                                    margin {safe_format_number(current_category.margin_amount,0)}")
                            except Exception as e:
                                logger.error(f"Error creating category for row {row}: {e}")
                                continue
                                
                        # Check if this is an item
                        elif (denominazione_val and current_category 
                              and not (codice_val and str(codice_val).startswith(IdentificationPatterns.GROUP_PREFIX))
                              and not (cod_val and len(str(cod_val).strip()) == IdentificationPatterns.CATEGORY_CODE_LENGTH)
                              and str(denominazione_val) != "DENOMINAZIONE"):  # Skip header row
                            
                            try:
                                # Extract all item fields including engineering costs
                                item = QuotationItem(
                                    position=str(self._safe_cell_value(row, ExcelColumns.POSITION, row)),
                                    code=str(codice_val) if codice_val else "",
                                    cod_listino=str(self._safe_cell_value(row, ExcelColumns.COD_LISTINO, "")),
                                    description=str(denominazione_val),
                                    quantity=float(self._safe_decimal(qta_val)),
                                    pricelist_unit_price=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.LIST_UNIT))),
                                    pricelist_total_price=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.LISTINO_TOTALE))),
                                    unit_cost=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.COSTO_UNITARIO))),
                                    total_cost=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.COSTO_TOTALE))),
                                    internal_code=str(self._safe_cell_value(row, ExcelColumns.COD_2, "")),
                                    priority_order=int(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.PRIORITY_ORDER, 0))),
                                    
                                    # Engineering and manufacturing costs
                                    utm_robot=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.UTM_ROBOT))),
                                    utm_robot_h=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.UTM_ROBOT_H))),
                                    utm_lgv=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.UTM_LGV))),
                                    utm_lgv_h=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.UTM_LGV_H))),
                                    utm_intra=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.UTM_INTRA))),
                                    utm_intra_h=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.UTM_INTRA_H))),
                                    utm_layout=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.UTM_LAYOUT))),
                                    utm_layout_h=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.UTM_LAYOUT_H))),
                                    
                                    # Engineering costs
                                    ute=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.UTE))),
                                    ute_h=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.UTE_H))),
                                    ba=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.BA))),
                                    ba_h=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.BA_H))),
                                    sw_pc=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.SW_PC))),
                                    sw_pc_h=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.SW_PC_H))),
                                    sw_plc=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.SW_PLC))),
                                    sw_plc_h=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.SW_PLC_H))),
                                    sw_lgv=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.SW_LGV))),
                                    sw_lgv_h=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.SW_LGV_H))),
                                    
                                    # Manufacturing costs  
                                    mtg_mec=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.MTG_MEC))),
                                    mtg_mec_h=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.MTG_MEC_H))),
                                    mtg_mec_intra=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.MTG_MEC_INTRA))),
                                    mtg_mec_intra_h=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.MTG_MEC_INTRA_H))),
                                    cab_ele=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.CAB_ELE))),
                                    cab_ele_h=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.CAB_ELE_H))),
                                    cab_ele_intra=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.CAB_ELE_INTRA))),
                                    cab_ele_intra_h=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.CAB_ELE_INTRA_H))),
                                    
                                    # Testing and field costs
                                    site=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.SITE))),
                                    site_h=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.SITE_H))),
                                    install=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.INSTALL))),
                                    install_h=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.INSTALL_H))),
                                    
                                    # Additional costs
                                    pm_cost=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.PM_COST))),
                                    pm_h=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.PM_H))),
                                    document=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.DOCUMENT))),
                                    document_h=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.DOCUMENT_H))),
                                    after_sales=float(self._safe_decimal(self._safe_cell_value(row, ExcelColumns.AFTER_SALES)))
                                )
                                
                                current_category.items.append(item)
                                logger.debug(f"Found item: {codice_val}")
                            except Exception as e:
                                logger.error(f"Error creating item for row {row}: {e}")
                                continue
                        
                    except Exception as e:
                        logger.error(f"Error processing row {row}: {e}")
                        continue
                        
            except Exception as e:
                error_msg = f"Error iterating through worksheet rows: {e}"
                logger.error(error_msg)
                raise Exception(error_msg) from e
            
            # Add the last group if exists
            try:
                if current_group:
                    product_groups.append(current_group)
            except Exception as e:
                logger.error(f"Error adding final group: {e}")
            
            # Create VA21 group for unmatched items  
            try:
                va21_group = ProductGroup(
                    group_id="VA21",
                    group_name="VA21",
                    quantity=1,
                    categories=[]
                )
                
                # Create a set of all WBE codes from existing categories
                existing_wbes = set()
                for group in product_groups:
                    for category in group.categories:
                        if category.wbe:
                            existing_wbes.add(category.wbe)

                # Create VA21 category and add unmatched items
                if self.va21_offers:
                    for wbe, offer_data in self.va21_offers.items():
                        try:
                            if wbe not in existing_wbes:
                                item = QuotationCategory(
                                    category_id=offer_data.get(VA21Columns.COD, ''),
                                    category_name=offer_data.get(VA21Columns.DESCRIPTION, ''),
                                    quantity=offer_data.get(VA21Columns.QUANTITY, 0.0),
                                    pricelist_subtotal=offer_data.get(VA21Columns.LISTINO_SUBTOTAL, 0.0),
                                    cost_subtotal=offer_data.get(VA21Columns.COST_SUBTOTAL, 0.0),
                                    total_cost=offer_data.get(VA21Columns.COST_SUBTOTAL, 0.0),
                                    margin_amount=offer_data.get(VA21Columns.OFFER_TOTAL, 0.0) - offer_data.get(VA21Columns.COST_SUBTOTAL, 0.0),
                                    margin_percentage=offer_data.get(VA21Columns.MARGIN_PERCENTAGE, 0.0),
                                    offer_price=offer_data.get(VA21Columns.OFFER_TOTAL, 0.0),
                                    wbe=wbe
                                )
                                
                                va21_group.categories.append(item)
                                logger.debug(f"Added VA21 category: {item.category_id} - \
                                    list {safe_format_number(item.pricelist_subtotal,0)} - \
                                    cost {safe_format_number(item.cost_subtotal,0)} - \
                                    offer {safe_format_number(item.offer_price,0)} - \
                                    margin {safe_format_number(item.margin_amount,0)}")
                        except Exception as e:
                            logger.error(f"Error creating VA21 category for WBE {wbe}: {e}")
                            continue
                
                # Only add VA21 group if it has categories
                if va21_group.categories:
                    product_groups.append(va21_group)
                    
            except Exception as e:
                logger.error(f"Error processing VA21 offers: {e}")
            
            logger.debug(f"Successfully extracted {len(product_groups)} product groups")
            return product_groups
            
        except RuntimeError as e:
            # Re-raise runtime errors
            raise e
        except Exception as e:
            error_msg = f"Unexpected error extracting product groups: {e}"
            logger.error(error_msg)
            raise Exception(error_msg) from e
    
    def calculate_totals(self, product_groups: List[ProductGroup], parameters: ProjectParameters) -> QuotationTotals:
        """Calculate total costs and fees"""
        try:
            logger.debug("Calculating totals")
            
            total_pricelist = Decimal("0.0")
            total_cost = Decimal("0.0")
            total_offer = Decimal("0.0")
            offer_margin = Decimal("0.0")
            offer_margin_percentage = Decimal("0.0")
            
            # Sum up costs from all categories
            try:
                for group in product_groups:
                    try:
                        for category in group.categories:
                            try:
                                # Use offer_price if available, otherwise pricelist_subtotal
                                category_pricelist = Decimal(str(category.pricelist_subtotal)) if category.pricelist_subtotal is not None else Decimal("0.0")
                                category_cost = Decimal(str(category.cost_subtotal)) if category.cost_subtotal is not None else Decimal("0.0")
                                category_offer = Decimal(str(category.offer_price)) if category.offer_price is not None else Decimal("0.0")
                                category_margin = Decimal(str(category.margin_amount)) if category.margin_amount is not None else Decimal("0.0")

                                total_pricelist += category_pricelist
                                total_cost += category_cost
                                total_offer += category_offer
                                offer_margin += category_margin
                                
                            except Exception as e:
                                logger.warning(f"Error processing category {category.category_id} in group {group.group_id}: {e}")
                                continue
                                
                    except Exception as e:
                        logger.warning(f"Error processing group {group.group_id}: {e}")
                        continue
                        
            except Exception as e:
                error_msg = f"Error iterating through product groups: {e}"
                logger.error(error_msg)
                raise Exception(error_msg) from e
            
            # Calculate margin percentage safely
            try:
                if total_cost != 0:
                    offer_margin_percentage = (offer_margin / total_cost) * 100
                else:
                    offer_margin_percentage = Decimal("0.0")
                    logger.warning("Total cost is zero, setting margin percentage to 0")
            except Exception as e:
                logger.warning(f"Error calculating margin percentage: {e}")
                offer_margin_percentage = Decimal("0.0")
            
            try:
                totals = QuotationTotals(
                    total_pricelist=float(self._round_decimal(total_pricelist)),
                    total_cost=float(self._round_decimal(total_cost)),
                    total_offer=float(self._round_decimal(total_offer)),
                    offer_margin=float(self._round_decimal(offer_margin)),
                    offer_margin_percentage=float(self._round_decimal(offer_margin_percentage))
                )
                logger.debug(f"Calculated totals: pricelist={totals.total_pricelist}, cost={totals.total_cost}, offer={totals.total_offer}")
                return totals
                
            except Exception as e:
                error_msg = f"Error creating QuotationTotals object: {e}"
                logger.error(error_msg)
                raise Exception(error_msg) from e
                
        except Exception as e:
            error_msg = f"Unexpected error calculating totals: {e}"
            logger.error(error_msg)
            raise Exception(error_msg) from e
    
    def close(self):
        """Close the workbook and release file handles"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self.ws = None

    def parse(self) -> IndustrialQuotation:
        """Main parsing method - returns IndustrialQuotation object directly"""
        try:
            logger.info(f"Starting direct parsing of Analisi Profittabilita file: {self.file_path}")
            
            try:
                self.load_workbook()
            except Exception as e:
                error_msg = f"Failed to load workbook: {e}"
                logger.error(error_msg)
                raise Exception(error_msg) from e
            
            # Extract all sections as model objects
            try:
                project_info = self.extract_project_info()
            except Exception as e:
                error_msg = f"Failed to extract project info: {e}"
                logger.error(error_msg)
                raise Exception(error_msg) from e
            
            try:
                product_groups = self.extract_product_groups()
            except Exception as e:
                error_msg = f"Failed to extract product groups: {e}"
                logger.error(error_msg)
                raise Exception(error_msg) from e
            
            try:
                totals = self.calculate_totals(product_groups, project_info.parameters)
            except Exception as e:
                error_msg = f"Failed to calculate totals: {e}"
                logger.error(error_msg)
                raise Exception(error_msg) from e
            
            # Create final IndustrialQuotation object
            try:
                quotation = IndustrialQuotation(
                    project=project_info,
                    product_groups=product_groups,
                    totals=totals,
                    source_file=str(self.file_path),
                    parser_type=ParserType.ANALISI_PROFITTABILITA_PARSER
                )
                
                logger.info(f"Direct parsing completed. Found {len(product_groups)} product groups")
                return quotation
                
            except Exception as e:
                error_msg = f"Failed to create IndustrialQuotation object: {e}"
                logger.error(error_msg)
                raise Exception(error_msg) from e
                
        except Exception as e:
            error_msg = f"Unexpected error during parsing: {e}"
            logger.error(error_msg)
            raise Exception(error_msg) from e
        finally:
            # Always close the workbook to release file handles
            try:
                self.close()
            except Exception as e:
                logger.warning(f"Error closing workbook: {e}")
    
    # Helper methods
    def _safe_cell_value(self, row: int, column: int, default: Any = None) -> Any:
        """Safely extract cell value by row and column index"""
        try:
            cell_value = self.ws.cell(row=row, column=column).value
            return cell_value if cell_value is not None else default
        except Exception:
            return default
    
    def _safe_decimal(self, value: Any, default: Decimal = None) -> Decimal:
        """Safely convert value to Decimal"""
        if value is None or value == "":
            return default if default is not None else Decimal("0.0")
        
        try:
            # Convert to string and clean up
            str_value = str(value).strip()
            
            # Handle empty strings after stripping
            if not str_value:
                return default if default is not None else Decimal("0.0")
            
            # Handle common non-numeric values
            if str_value.lower() in ['n/a', 'na', 'null', 'none', '-', '']:
                return default if default is not None else Decimal("0.0")
                
            # Remove currency symbols and common formatting
            str_value = str_value.replace('€', '').replace('$', '').replace(',', '').strip()
            
            # Handle percentage notation
            if str_value.endswith('%'):
                str_value = str_value[:-1]
                return Decimal(str_value) / Decimal("100")
            
            return Decimal(str_value)
            
        except (ValueError, TypeError, decimal.InvalidOperation) as e:
            logger.debug(f"Could not convert '{value}' to Decimal: {e}")
            return default if default is not None else Decimal("0.0")
    
    def _safe_int(self, value: Any, default: int = 0) -> int:
        """Safely convert value to int"""
        if value is None:
            return default
        try:
            return int(float(value))  # Convert via float to handle decimal values
        except (ValueError, TypeError):
            return default
    
    def _round_decimal(self, value: Decimal) -> Decimal:
        """Round decimal to 2 places"""
        return value.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    
    def _find_latest_va21_sheet(self) -> Optional[str]:
        """Find the latest VA21 sheet in the workbook"""
        try:
            if not self.workbook:
                logger.warning("Workbook not loaded when trying to find VA21 sheet")
                return None
                
            va21_sheets = [name for name in self.workbook.sheetnames 
                           if name.startswith(IdentificationPatterns.VA21_SHEET_PREFIX)]
            
            if not va21_sheets:
                logger.debug("No VA21 sheets found in workbook")
                return None
            
            # Sort sheets to get the latest one
            va21_sheets.sort(reverse=True)
            logger.debug(f"Found VA21 sheets: {va21_sheets}, using: {va21_sheets[0]}")
            return va21_sheets[0]
            
        except Exception as e:
            error_msg = f"Error finding VA21 sheet: {e}"
            logger.error(error_msg)
            raise Exception(error_msg) from e
    
    def _convert_wbe_us_to_it(self, wbe_us: str) -> str:
        """Convert US WBE format to IT format"""
        try:
            if not wbe_us:
                return wbe_us
                
            if wbe_us.endswith(IdentificationPatterns.WBE_US_SUFFIX):
                result = wbe_us.replace(IdentificationPatterns.WBE_US_SUFFIX, 
                                      IdentificationPatterns.WBE_IT_SUFFIX)
                logger.debug(f"Converted WBE from {wbe_us} to {result}")
                return result
            return wbe_us
            
        except Exception as e:
            error_msg = f"Error converting WBE format from '{wbe_us}': {e}"
            logger.error(error_msg)
            raise Exception(error_msg) from e

# =============================================================================
# CONVENIENCE FUNCTIONS
# =============================================================================

def parse_analisi_profittabilita_direct(file_path: str, output_path: Optional[str] = None) -> IndustrialQuotation:
    """
    Parse Analisi Profittabilita Excel file directly to IndustrialQuotation object
    
    Args:
        file_path: Path to the Excel file
        output_path: Optional path to save JSON output
        
    Returns:
        IndustrialQuotation: Validated quotation model instance
    """
    parser = None
    try:
        logger.info(f"Starting parsing process for file: {file_path}")
        
        try:
            parser = DirectAnalisiProfittabilitaParser(file_path)
        except Exception as e:
            error_msg = f"Failed to initialize parser for file {file_path}: {e}"
            logger.error(error_msg)
            raise Exception(error_msg) from e
        
        try:
            quotation = parser.parse()
        except Exception as e:
            error_msg = f"Failed to parse file {file_path}: {e}"
            logger.error(error_msg)
            raise Exception(error_msg) from e
        
        if output_path:
            try:
                quotation.save_json(output_path)
                logger.info(f"JSON output saved to {output_path}")
            except Exception as e:
                error_msg = f"Failed to save JSON output to {output_path}: {e}"
                logger.error(error_msg)
                raise Exception(error_msg) from e
        
        logger.info("Parsing process completed successfully")
        return quotation
        
    except Exception as e:
        error_msg = f"Unexpected error in parse_analisi_profittabilita_direct: {e}"
        logger.error(error_msg)
        raise Exception(error_msg) from e
    finally:
        if parser:
            try:
                parser.close()
            except Exception as e:
                logger.warning(f"Error closing parser: {e}")

def validate_analisi_profittabilita_file(file_path: str) -> Dict[str, Any]:
    """
    Validate Analisi Profittabilita file and return validation results
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        Dictionary with validation results
    """
    parser = None
    try:
        logger.info(f"Starting validation process for file: {file_path}")
        
        try:
            parser = DirectAnalisiProfittabilitaParser(file_path)
        except Exception as e:
            error_msg = f"Failed to initialize parser for validation of {file_path}: {e}"
            logger.error(error_msg)
            return {
                "is_valid": False,
                "validation_results": {},
                "summary_stats": {},
                "errors": [error_msg]
            }
        
        try:
            quotation = parser.parse()
        except Exception as e:
            error_msg = f"Failed to parse file during validation {file_path}: {e}"
            logger.error(error_msg)
            return {
                "is_valid": False,
                "validation_results": {},
                "summary_stats": {},
                "errors": [error_msg]
            }
        
        # Run validation checks
        try:
            validation_results = quotation.validate_totals_consistency()
            summary_stats = quotation.get_summary_stats()
            
            logger.info("Validation process completed successfully")
            return {
                "is_valid": True,
                "validation_results": validation_results,
                "summary_stats": summary_stats,
                "errors": []
            }
            
        except Exception as e:
            error_msg = f"Failed to run validation checks: {e}"
            logger.error(error_msg)
            return {
                "is_valid": False,
                "validation_results": {},
                "summary_stats": {},
                "errors": [error_msg]
            }
        
    except Exception as e:
        error_msg = f"Unexpected error during validation: {e}"
        logger.error(error_msg)
        return {
            "is_valid": False,
            "validation_results": {},
            "summary_stats": {},
            "errors": [error_msg]
        }
    finally:
        if parser:
            try:
                parser.close()
            except Exception as e:
                logger.warning(f"Error closing parser during validation: {e}")

if __name__ == "__main__":
    # Example usage
    input_file = "input/test_ap.xlsm"
    output_file = "output/quotation_ap_direct.json"
    
    try:
        quotation = parse_analisi_profittabilita_direct(input_file, output_file)
        stats = quotation.get_summary_stats()
        
        print("✅ Direct parsing completed successfully!")
        print(f"📋 Project ID: {stats['project_id']}")
        print(f"🏢 Customer: {quotation.project.customer or 'Not specified'}")
        print(f"📦 Product groups: {stats['total_groups']}")
        print(f"📂 Categories: {stats['total_categories']}")
        print(f"📄 Items: {stats['total_items']}")
        print(f"💵 Currency: {quotation.project.parameters.currency}")
        print(f"💰 Total pricelist: €{quotation.totals.total_pricelist:,.2f}")
        print(f"💰 Total cost: €{quotation.totals.total_cost:,.2f}")
        print(f"💰 Total offer: €{quotation.totals.total_offer:,.2f}")
        print(f"💰 Offer margin: €{quotation.totals.offer_margin:,.2f}")
        print(f"💰 Offer margin percentage: {quotation.totals.offer_margin_percentage:.2f}%")
        
        # Validate consistency
        validation = quotation.validate_totals_consistency()
        print(f"\n🔍 Validation results:")
        for check, is_valid in validation.items():
            status = "✅" if is_valid else "❌"
            print(f"   {status} {check}")
        
    except Exception as e:
        logger.error(f"Error parsing Excel file: {e}")
        raise 