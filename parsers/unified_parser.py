"""
Unified Parser Interface for Industrial Equipment Quotations
Automatically detects file type and uses the appropriate parser to return validated models
"""

import logging
import os
from typing import Optional, Union
from pathlib import Path

# Import parsers and models
import sys

from parsers.pre_file_parser_direct import DirectPreFileParser
from parsers.analisi_profittabilita_parser_direct import DirectAnalisiProfittabilitaParser
sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from models import IndustrialQuotation

logger = logging.getLogger(__name__)

class UnifiedQuotationParser:
    """
    Unified parser interface that automatically detects quotation file types
    and uses the appropriate parser to return validated IndustrialQuotation models.
    """
    
    # File type detection patterns
    PRE_FILE_INDICATORS = [
        "PRE_",
        "pre_",
        "OFFER1",
        "offer1"
    ]
    
    ANALISI_PROFITTABILITA_INDICATORS = [
        "Analisi profitabilita",
        "analisi_profittabilita", 
        "profitability",
        "Tabella riassuntiva SAP",
        "NEW_OFFER1"
    ]
    
    def __init__(self, file_path: str):
        """
        Initialize unified parser with file path
        
        Args:
            file_path: Path to the Excel quotation file
        """
        self.file_path = file_path
        self.file_name = Path(file_path).name
        self.detected_type = self._detect_file_type()
        
        logger.info(f"Detected file type: {self.detected_type} for file: {self.file_name}")
    
    def _detect_file_type(self) -> str:
        """
        Detect the type of quotation file based on the presence of the 'NEW_OFFER1' sheet.
        
        Returns:
            File type: 'pre' or 'analisi_profittabilita'
        """
        from openpyxl import load_workbook

        try:
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
            try:
                sheet_names = [sheet_name.strip() for sheet_name in wb.sheetnames]
                if "NEW_OFFER1" in sheet_names:
                    logger.info(f"'NEW_OFFER1' sheet found in {self.file_name}: detected as analisi_profittabilita")
                    return 'analisi_profittabilita'
                else:
                    logger.info(f"'NEW_OFFER1' sheet not found in {self.file_name}: detected as pre")
                    return 'pre'
            finally:
                wb.close()
        except Exception as e:
            logger.warning(f"Could not open file {self.file_name} to detect sheet names: {e}. Defaulting to analisi_profittabilita.")
            return 'analisi_profittabilita'
    
    def parse(self) -> IndustrialQuotation:
        """
        Parse the quotation file using the appropriate parser
        
        Returns:
            IndustrialQuotation: Validated quotation model instance
        """
        logger.info(f"Parsing {self.detected_type} file: {self.file_name}")
        
        try:
            if self.detected_type == 'pre':
                parser = DirectPreFileParser(self.file_path)
                return parser.parse()
            else:  # analisi_profittabilita
                parser = DirectAnalisiProfittabilitaParser(self.file_path)
                return parser.parse()
                
        except Exception as e:
            logger.error(f"Failed to parse {self.file_name} as {self.detected_type}: {e}")
            
            # Try the other parser as fallback
            fallback_type = 'analisi_profittabilita' if self.detected_type == 'pre' else 'pre'
            logger.info(f"Attempting fallback parsing as {fallback_type}")
            
            try:
                if fallback_type == 'pre':
                    parser = DirectPreFileParser(self.file_path)
                    return parser.parse()
                else:
                    parser = DirectAnalisiProfittabilitaParser(self.file_path)
                    return parser.parse()
            except Exception as fallback_error:
                logger.error(f"Fallback parsing also failed: {fallback_error}")
                raise Exception(f"Could not parse {self.file_name} with either parser. Original error: {e}, Fallback error: {fallback_error}")
    
    def parse_and_save(self, output_path: str) -> IndustrialQuotation:
        """
        Parse the quotation file and save the result to JSON
        
        Args:
            output_path: Path to save the JSON output
            
        Returns:
            IndustrialQuotation: Validated quotation model instance
        """
        quotation = self.parse()
        quotation.save_json(output_path)
        logger.info(f"Saved parsed quotation to: {output_path}")
        return quotation
    
    def get_parser_recommendations(self) -> dict:
        """
        Get recommendations for which parser to use based on file analysis
        
        Returns:
            Dictionary with parser recommendations and confidence scores
        """
        file_name_lower = self.file_name.lower()
        
        pre_score = 0
        ap_score = 0
        
        # Score based on filename indicators
        for indicator in self.PRE_FILE_INDICATORS:
            if indicator.lower() in file_name_lower:
                pre_score += 10
        
        for indicator in self.ANALISI_PROFITTABILITA_INDICATORS:
            if indicator.lower() in file_name_lower:
                ap_score += 10
        
        # Additional scoring based on file patterns
        if 'pre' in file_name_lower and 'only' in file_name_lower:
            pre_score += 15
        
        if 'tabella' in file_name_lower or 'sap' in file_name_lower:
            ap_score += 15
        
        if 'va21' in file_name_lower:
            ap_score += 20
        
        # Normalize scores
        total_score = pre_score + ap_score
        if total_score > 0:
            pre_confidence = pre_score / total_score * 100
            ap_confidence = ap_score / total_score * 100
        else:
            pre_confidence = 50
            ap_confidence = 50
        
        recommended_parser = 'pre' if pre_score > ap_score else 'analisi_profittabilita'
        
        return {
            'recommended_parser': recommended_parser,
            'pre_confidence': round(pre_confidence, 1),
            'analisi_profittabilita_confidence': round(ap_confidence, 1),
            'detected_type': self.detected_type,
            'pre_score': pre_score,
            'ap_score': ap_score
        }

def parse_quotation_file(file_path: str, output_path: Optional[str] = None, 
                        force_parser: Optional[str] = None) -> IndustrialQuotation:
    """
    Unified function to parse any quotation file to IndustrialQuotation model
    
    Args:
        file_path: Path to the Excel quotation file
        output_path: Optional path to save JSON output
        force_parser: Optional parser type to force ('pre' or 'analisi_profittabilita')
        
    Returns:
        IndustrialQuotation: Validated quotation model instance
    """
    if force_parser:
        logger.info(f"Forcing parser type: {force_parser}")
        if force_parser == 'pre':
            from parsers.pre_file_parser_direct import parse_pre_file_direct
            return parse_pre_file_direct(file_path, output_path)
        elif force_parser == 'analisi_profittabilita':
            from parsers.analisi_profittabilita_parser_direct import parse_analisi_profittabilita_direct
            return parse_analisi_profittabilita_direct(file_path, output_path)
        else:
            raise ValueError(f"Invalid forced parser type: {force_parser}")
    
    # Use automatic detection
    parser = UnifiedQuotationParser(file_path)
    
    if output_path:
        return parser.parse_and_save(output_path)
    else:
        return parser.parse()

def analyze_quotation_file(file_path: str) -> dict:
    """
    Analyze a quotation file and provide parsing recommendations
    
    Args:
        file_path: Path to the Excel quotation file
        
    Returns:
        Dictionary with file analysis and parser recommendations
    """
    parser = UnifiedQuotationParser(file_path)
    recommendations = parser.get_parser_recommendations()
    
    file_info = {
        'file_path': file_path,
        'file_name': parser.file_name,
        'file_size': os.path.getsize(file_path) if os.path.exists(file_path) else None,
        'exists': os.path.exists(file_path)
    }
    
    return {
        'file_info': file_info,
        'parser_recommendations': recommendations
    } 