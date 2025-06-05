"""
Excel to JSON Parser for Industrial Equipment Quotations
Converts Excel files with specific format to structured JSON according to schema
"""

import json
import logging
from typing import Dict, List, Optional, Any
import pandas as pd
from openpyxl import load_workbook

# Import unified models and field mappings
import sys
import os
sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from models import IndustrialQuotation, FieldMapper, ParserType

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# =============================================================================
# CONSTANTS
# =============================================================================

# Excel Column Indices (1-based)
class ExcelColumns:
    COD = 1
    CODICE = 3
    DENOMINAZIONE = 4
    QTA = 5
    LIST_UNIT = 6
    LISTINO = 7
    SUB_TOT_LISTINO = 8
    SUB_TOT_CODICE = 9
    TOTALE = 10
    GRUPPI = 11
    TOTALE_OFFERTA = 12
    VALUTA = 13
    TOTALE_CODICE = 14
    TOTALE_SCONTATO = 15
    NOTE = 16
    COD_LISTINO = 17
    TOTAL_DISCOUNTED = 18
    COSTO_UNITARIO = 19
    COSTO = 20
    SUB_TOT_COSTO = 21
    TOTALE_COSTO = 22

# Excel Row Constants
class ExcelRows:
    HEADER_ROW = 17
    DATA_START_ROW = 18
    PROJECT_INFO_START = 1
    PROJECT_INFO_END = 16

# Identification Patterns
class IdentificationPatterns:
    GROUP_PREFIX = 'TXT-'
    CATEGORY_CODE_LENGTH = 4
    HEADER_CODE = 'COD'
# Project Information Cell Positions (row, column)
class ProjectInfoCells:
    PROJECT_ID = (1, 1)  # Row 1, Column A
    CUSTOMER = (3, 7)    # Row 3, Column G
    DOC_PERCENTAGE = (8, 2)     # Row 8, Column B
    PM_PERCENTAGE = (9, 2)      # Row 8, Column D
    FINANCIAL_COSTS = (10, 2)    # Row 9, Column D
    CURRENCY = (11, 2)          # Row 11, Column D
    EXCHANGE_RATE = (12, 2)     # Row 12, Column D
    WASTE_DISPOSAL = (13, 2)    # Row 13, Column D
    WARRANTY_PERCENTAGE = (8, 11) # Row 14, Column D

# JSON Field Names
class JsonFields:
    # Project fields
    PROJECT = "project"
    ID = "id"
    CUSTOMER = "customer"
    PARAMETERS = "parameters"
    SALES_INFO = "sales_info"
    
    # Parameter fields
    DOC_PERCENTAGE = "doc_percentage"
    PM_PERCENTAGE = "pm_percentage"
    FINANCIAL_COSTS = "financial_costs"
    CURRENCY = "currency"
    EXCHANGE_RATE = "exchange_rate"
    WASTE_DISPOSAL = "waste_disposal"
    WARRANTY_PERCENTAGE = "warranty_percentage"
    IS_24H_SERVICE = "is_24h_service"
    
    # Sales info fields
    AREA_MANAGER = "area_manager"
    AGENT = "agent"
    COMMISSION_PERCENTAGE = "commission_percentage"
    AUTHOR = "author"
    
    # Product group fields
    PRODUCT_GROUPS = "product_groups"
    GROUP_ID = "group_id"
    GROUP_NAME = "group_name"
    QUANTITY = "quantity"
    CATEGORIES = "categories"
    
    # Category fields
    CATEGORY_ID = "category_id"
    CATEGORY_NAME = "category_name"
    ITEMS = "items"
    SUBTOTAL_LISTINO = "subtotal_listino"
    SUBTOTAL_CODICE = "subtotal_codice"
    TOTAL = "total"
    GROUPS = "groups"
    TOTAL_OFFER = "total_offer"
    TOTAL_OFFER_CURRENCY = "total_offer_currency"
    TOTAL_CODE_CURRENCY = "total_code_currency"
    TOTAL_DISCOUNTED_CURRENCY = "total_discounted_currency"
    NOTES = "notes"
    PRICELIST_CODE = "pricelist_code"
    TOTAL_DISCOUNTED = "total_discounted"
    UNIT_COST = "unit_cost"
    UNIT_TOTAL_COST = "unit_total_cost"
    SUBTOTAL_COST = "subtotal_cost"
    TOTAL_COST = "total_cost"
    
    # Item fields
    POSITION = "position"
    CODE = "code"
    DESCRIPTION = "description"
    PRICELIST_UNIT_PRICE = "pricelist_unit_price"
    PRICELIST_TOTAL_PRICE = "pricelist_total_price"
    
    # Totals fields
    TOTALS = "totals"
    EQUIPMENT_TOTAL = "equipment_total"
    INSTALLATION_TOTAL = "installation_total"
    SUBTOTAL = "subtotal"
    DOC_FEE = "doc_fee"
    PM_FEE = "pm_fee"
    WARRANTY_FEE = "warranty_fee"
    GRAND_TOTAL = "grand_total"

# Calculation Constants
class CalculationConstants:
    DEFAULT_QUANTITY = 1
    DEFAULT_FLOAT = 0.0
    DEFAULT_INT = 0

# Log Messages
class LogMessages:
    PARSING_START = "Starting to parse {}"
    WORKBOOK_LOADED = "Loaded workbook with {} rows and {} columns"
    GROUP_FOUND = "Found group: {}"
    CATEGORY_FOUND = "Found category: {}"
    ITEM_FOUND = "Found item: {}"
    PARSING_COMPLETED = "Parsing completed. Found {} product groups"
    JSON_SAVED = "JSON output saved to {}"

# Error Messages
class ErrorMessages:
    FILE_NOT_FOUND = "File not found: {}"
    INVALID_WORKBOOK = "Unable to load workbook: {}"
    MISSING_SHEET = "Sheet '{}' not found in workbook"
    INVALID_DATA_FORMAT = "Invalid data format in row {}"

# =============================================================================
# MAIN PARSER CLASS
# =============================================================================

class PreFileParser:
    """Parser for converting Excel quotations to JSON format"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.ws = None
        
    def load_workbook(self):
        """Load the Excel workbook"""
        try:
            self.workbook = load_workbook(self.file_path, data_only=True)
            # Use the first worksheet
            self.ws = self.workbook['OFFER1']
            logger.info(LogMessages.WORKBOOK_LOADED.format(self.ws.max_row, self.ws.max_column))
        except FileNotFoundError:
            raise FileNotFoundError(ErrorMessages.FILE_NOT_FOUND.format(self.file_path))
        except Exception as e:
            raise Exception(ErrorMessages.INVALID_WORKBOOK.format(str(e)))
    
    def extract_project_info(self) -> Dict[str, Any]:
        """Extract project information from the Excel file"""
        # Extract basic project info
        project_id_cell = self.ws.cell(row=ProjectInfoCells.PROJECT_ID[0], column=ProjectInfoCells.PROJECT_ID[1])
        customer_cell = self.ws.cell(row=ProjectInfoCells.CUSTOMER[0], column=ProjectInfoCells.CUSTOMER[1])
        
        project_id = self._extract_after_colon(project_id_cell.value) if project_id_cell.value else ""
        customer = self._extract_after_colon(customer_cell.value) if customer_cell.value else ""
        
        # Extract parameters
        doc_cell = self.ws.cell(row=ProjectInfoCells.DOC_PERCENTAGE[0], column=ProjectInfoCells.DOC_PERCENTAGE[1])
        pm_cell = self.ws.cell(row=ProjectInfoCells.PM_PERCENTAGE[0], column=ProjectInfoCells.PM_PERCENTAGE[1])
        financial_cell = self.ws.cell(row=ProjectInfoCells.FINANCIAL_COSTS[0], column=ProjectInfoCells.FINANCIAL_COSTS[1])
        currency_cell = self.ws.cell(row=ProjectInfoCells.CURRENCY[0], column=ProjectInfoCells.CURRENCY[1])
        exchange_cell = self.ws.cell(row=ProjectInfoCells.EXCHANGE_RATE[0], column=ProjectInfoCells.EXCHANGE_RATE[1])
        waste_cell = self.ws.cell(row=ProjectInfoCells.WASTE_DISPOSAL[0], column=ProjectInfoCells.WASTE_DISPOSAL[1])
        warranty_cell = self.ws.cell(row=ProjectInfoCells.WARRANTY_PERCENTAGE[0], column=ProjectInfoCells.WARRANTY_PERCENTAGE[1])
        
        return {
            JsonFields.ID: project_id,
            JsonFields.CUSTOMER: customer,
            JsonFields.PARAMETERS: {
                JsonFields.DOC_PERCENTAGE: self._safe_float(self._extract_after_colon(doc_cell.value)),
                JsonFields.PM_PERCENTAGE: self._safe_float(self._extract_after_colon(pm_cell.value)),
                JsonFields.FINANCIAL_COSTS: self._safe_float(self._extract_after_colon(financial_cell.value)),
                JsonFields.CURRENCY: self._extract_after_colon(currency_cell.value) if currency_cell.value else "",
                JsonFields.EXCHANGE_RATE: self._safe_float(self._extract_after_colon(exchange_cell.value)),
                JsonFields.WASTE_DISPOSAL: self._safe_float(self._extract_after_colon(waste_cell.value)),
                JsonFields.WARRANTY_PERCENTAGE: self._safe_float(self._extract_after_colon(warranty_cell.value)),
                JsonFields.IS_24H_SERVICE: False  # Default value, could be extracted if present
            },
            JsonFields.SALES_INFO: {
                JsonFields.AREA_MANAGER: None,  # Not present in current Excel format
                JsonFields.AGENT: None,  # Not present in current Excel format
                JsonFields.COMMISSION_PERCENTAGE: CalculationConstants.DEFAULT_FLOAT,
                JsonFields.AUTHOR: None  # Not present in current Excel format
            }
        }
    
    def extract_product_groups(self) -> List[Dict[str, Any]]:
        """Extract product groups, categories, and items"""
        product_groups = []
        current_group = None
        current_category = None
        
        # Start from data start row
        for row in range(ExcelRows.DATA_START_ROW, self.ws.max_row + 1):
            cod_val = self.ws.cell(row=row, column=ExcelColumns.COD).value
            codice_val = self.ws.cell(row=row, column=ExcelColumns.CODICE).value
            denominazione_val = self.ws.cell(row=row, column=ExcelColumns.DENOMINAZIONE).value
            qta_val = self.ws.cell(row=row, column=ExcelColumns.QTA).value
            listino_val = self.ws.cell(row=row, column=ExcelColumns.LIST_UNIT).value
            listino_tot_val = self.ws.cell(row=row, column=ExcelColumns.LISTINO).value
            sub_tot_listino_val = self.ws.cell(row=row, column=ExcelColumns.SUB_TOT_LISTINO).value
            sub_tot_codice_val = self.ws.cell(row=row, column=ExcelColumns.SUB_TOT_CODICE).value
            tot_val = self.ws.cell(row=row, column=ExcelColumns.TOTALE).value
            gruppi_val = self.ws.cell(row=row, column=ExcelColumns.GRUPPI).value
            tot_offer_val = self.ws.cell(row=row, column=ExcelColumns.TOTALE_OFFERTA).value
            valuta_val = self.ws.cell(row=row, column=ExcelColumns.VALUTA).value
            totale_codice_val = self.ws.cell(row=row, column=ExcelColumns.TOTALE_CODICE).value
            totale_scontato_val = self.ws.cell(row=row, column=ExcelColumns.TOTALE_SCONTATO).value
            note_val = self.ws.cell(row=row, column=ExcelColumns.NOTE).value
            cod_listino_val = self.ws.cell(row=row, column=ExcelColumns.COD_LISTINO).value  
            totale_scontato_val = self.ws.cell(row=row, column=ExcelColumns.TOTAL_DISCOUNTED).value
            costo_unitario_val = self.ws.cell(row=row, column=ExcelColumns.COSTO_UNITARIO).value
            costo_val = self.ws.cell(row=row, column=ExcelColumns.COSTO).value
            sub_tot_costo_val = self.ws.cell(row=row, column=ExcelColumns.SUB_TOT_COSTO).value
            tot_costo_val = self.ws.cell(row=row, column=ExcelColumns.TOTALE_COSTO).value
            


            # Check if this is a group header
            if codice_val and str(codice_val).startswith(IdentificationPatterns.GROUP_PREFIX):
                # Save previous group if exists
                if current_group:
                    product_groups.append(current_group)
                
                # Start new group
                current_group = {
                    JsonFields.GROUP_ID: str(codice_val),
                    JsonFields.GROUP_NAME: str(denominazione_val) if denominazione_val else "",
                    JsonFields.QUANTITY: self._safe_int(qta_val, CalculationConstants.DEFAULT_QUANTITY),
                    JsonFields.CATEGORIES: []
                }
                current_category = None
                logger.info(LogMessages.GROUP_FOUND.format(codice_val))
                
            # Check if this is a category
            elif cod_val and len(str(cod_val).strip()) == IdentificationPatterns.CATEGORY_CODE_LENGTH and current_group:
                current_category = {
                    JsonFields.CATEGORY_ID: str(cod_val),
                    JsonFields.CODE: str(codice_val),
                    JsonFields.CATEGORY_NAME: str(denominazione_val) if denominazione_val else "",
                    JsonFields.ITEMS: [],
                    JsonFields.SUBTOTAL_LISTINO: self._safe_float(sub_tot_listino_val),
                    JsonFields.SUBTOTAL_CODICE: self._safe_float(sub_tot_codice_val),
                    JsonFields.TOTAL: self._safe_float(tot_val),
                    JsonFields.GROUPS: self._safe_float(gruppi_val),
                    JsonFields.TOTAL_OFFER: self._safe_float(tot_offer_val),
                    JsonFields.TOTAL_OFFER_CURRENCY: self._safe_float(valuta_val),
                    JsonFields.TOTAL_CODE_CURRENCY: self._safe_float(valuta_val),
                    JsonFields.TOTAL_DISCOUNTED_CURRENCY: self._safe_float(valuta_val),
                    JsonFields.NOTES: self._safe_float(note_val),
                    JsonFields.PRICELIST_CODE: self._safe_float(cod_listino_val),
                    JsonFields.TOTAL_DISCOUNTED: self._safe_float(totale_scontato_val),
                    JsonFields.UNIT_COST: self._safe_float(costo_unitario_val),
                    JsonFields.TOTAL_COST: self._safe_float(costo_val),
                    JsonFields.SUBTOTAL_COST: self._safe_float(sub_tot_costo_val),
                    JsonFields.TOTAL_COST: self._safe_float(tot_costo_val),
                    
                    
                }
                current_group[JsonFields.CATEGORIES].append(current_category)
                logger.info(LogMessages.CATEGORY_FOUND.format(cod_val))
                
            # Check if this is an item
            elif codice_val and denominazione_val and current_category \
                and not str(codice_val).startswith(IdentificationPatterns.GROUP_PREFIX) \
                and not str(codice_val).startswith(IdentificationPatterns.HEADER_CODE) :
                item = {
                    JsonFields.POSITION: str(row),
                    JsonFields.CODE: str(codice_val),
                    JsonFields.DESCRIPTION: str(denominazione_val),
                    JsonFields.QUANTITY: self._safe_float(qta_val),
                    JsonFields.PRICELIST_UNIT_PRICE: self._safe_float(listino_val),
                    JsonFields.PRICELIST_TOTAL_PRICE: self._safe_float(listino_tot_val),
                    JsonFields.NOTES: self._safe_float(note_val),
                    JsonFields.PRICELIST_CODE: self._safe_float(cod_listino_val),
                    JsonFields.UNIT_COST: self._safe_float(costo_unitario_val),
                    JsonFields.TOTAL_COST: self._safe_float(costo_val),

                }
                
                current_category[JsonFields.ITEMS].append(item)
                logger.debug(LogMessages.ITEM_FOUND.format(codice_val))
        
        # Add the last group if exists
        if current_group:
            product_groups.append(current_group)
        
        # Calculate category totals
        for group in product_groups:
            for category in group[JsonFields.CATEGORIES]:
                if not category[JsonFields.SUBTOTAL_LISTINO]:
                    category[JsonFields.SUBTOTAL_LISTINO] = sum(item[JsonFields.PRICELIST_TOTAL_PRICE] for item in category[JsonFields.ITEMS])
                if not category[JsonFields.SUBTOTAL_CODICE]:
                    category[JsonFields.SUBTOTAL_CODICE] = category[JsonFields.SUBTOTAL_LISTINO]
                if not category[JsonFields.TOTAL]:
                    category[JsonFields.TOTAL] = category[JsonFields.SUBTOTAL_CODICE]
        
        return product_groups
    
    def calculate_totals(self, product_groups: List[Dict[str, Any]], parameters: Dict[str, Any]) -> Dict[str, Any]:
        """Calculate total costs and fees"""
        equipment_total = CalculationConstants.DEFAULT_FLOAT
        installation_total = CalculationConstants.DEFAULT_FLOAT
        
        # Sum up equipment costs from all categories
        equipment_total = sum(category.get(JsonFields.TOTAL_OFFER, 0) 
                               for group in product_groups 
                               for category in group[JsonFields.CATEGORIES]
                               if not str(category.get(JsonFields.CATEGORY_ID, '')).startswith('E'))
        
        # Calculate installation costs
        installation_total = sum(category.get(JsonFields.TOTAL_OFFER, 0) 
                               for group in product_groups 
                               for category in group[JsonFields.CATEGORIES]
                               if str(category.get(JsonFields.CATEGORY_ID, '')).startswith('E'))
        
        # Calculate subtotal
        subtotal = equipment_total + installation_total
        
        # Calculate fees
        doc_fee = subtotal * parameters[JsonFields.DOC_PERCENTAGE]
        pm_fee = subtotal * parameters[JsonFields.PM_PERCENTAGE]
        warranty_fee = subtotal * parameters.get(JsonFields.WARRANTY_PERCENTAGE, CalculationConstants.DEFAULT_FLOAT)
        
        # Calculate grand total
        grand_total = subtotal + doc_fee + pm_fee + warranty_fee
        
        return {
            JsonFields.EQUIPMENT_TOTAL: round(equipment_total, 2),
            JsonFields.INSTALLATION_TOTAL: round(installation_total, 2),
            JsonFields.SUBTOTAL: round(subtotal, 2),
            JsonFields.DOC_FEE: round(doc_fee, 2),
            JsonFields.PM_FEE: round(pm_fee, 2),
            JsonFields.WARRANTY_FEE: round(warranty_fee, 2),
            JsonFields.GRAND_TOTAL: round(grand_total, 2)
        }
    
    def parse(self) -> Dict[str, Any]:
        """Main parsing method"""
        logger.info(LogMessages.PARSING_START.format(self.file_path))
        
        self.load_workbook()
        
        # Extract all sections
        project_info = self.extract_project_info()
        product_groups = self.extract_product_groups()
        totals = self.calculate_totals(product_groups, project_info[JsonFields.PARAMETERS])
        
        # Build final structure
        result = {
            JsonFields.PROJECT: project_info,
            JsonFields.PRODUCT_GROUPS: product_groups,
            JsonFields.TOTALS: totals
        }
        
        logger.info(LogMessages.PARSING_COMPLETED.format(len(product_groups)))
        return result
    
    def parse_to_model(self) -> IndustrialQuotation:
        """
        Parse Excel file directly to IndustrialQuotation model with validation
        
        Returns:
            IndustrialQuotation: Validated quotation model instance
        """
        logger.info(f"Parsing PRE file to IndustrialQuotation model: {self.file_path}")
        
        # Get raw parser data
        raw_data = self.parse()
        
        # Convert to unified model format using field mapper
        converted_data = FieldMapper.convert_pre_parser_dict(raw_data)
        
        # Create and validate model instance
        quotation = IndustrialQuotation.from_parser_dict(
            converted_data, 
            source_file=self.file_path,
            parser_type=ParserType.PRE_FILE_PARSER
        )
        
        logger.info(f"Successfully created IndustrialQuotation model with {quotation.get_summary_stats()['total_items']} items")
        return quotation
    
    def _safe_float(self, value: Any, default: float = CalculationConstants.DEFAULT_FLOAT) -> float:
        """Safely convert value to float"""
        if value is None:
            return default
        try:
            return float(value)
        except (ValueError, TypeError):
            return default
    
    def _safe_int(self, value: Any, default: int = CalculationConstants.DEFAULT_INT) -> int:
        """Safely convert value to int"""
        if value is None:
            return default
        try:
            return int(value)
        except (ValueError, TypeError):
            return default
    
    def _extract_after_colon(self, value: Any) -> str:
        """Extract text after colon, or return the value as string"""
        if value is None:
            return ""
        
        str_value = str(value)
        if ':' in str_value:
            return str_value.split(':', 1)[1].strip()
        return str_value.strip()

def parse_pre_to_json(file_path: str, output_path: Optional[str] = None) -> Dict[str, Any]:
    """
    Main function to parse Excel file and optionally save to JSON
    
    Args:
        file_path: Path to the Excel file
        output_path: Optional path to save JSON output
        
    Returns:
        Dictionary containing the parsed data
    """
    parser = PreFileParser(file_path)
    result = parser.parse()
    
    if output_path:
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, indent=2, ensure_ascii=False)
        logger.info(LogMessages.JSON_SAVED.format(output_path))
    
    return result

def parse_pre_to_model(file_path: str, output_path: Optional[str] = None) -> IndustrialQuotation:
    """
    Main function to parse Excel file to IndustrialQuotation model with validation
    
    Args:
        file_path: Path to the Excel file
        output_path: Optional path to save JSON output using model serialization
        
    Returns:
        IndustrialQuotation: Validated quotation model instance
    """
    parser = PreFileParser(file_path)
    quotation = parser.parse_to_model()
    
    if output_path:
        quotation.save_json(output_path)
    
    return quotation


if __name__ == "__main__":
    # Example usage
    input_file = "input/PRE_ONLY_OFFER1.xlsx"
    output_file = "output/quotation.json"
    
    try:
        result = parse_pre_to_json(input_file, output_file)
        print("Parsing completed successfully!")
        print(f"Project ID: {result['project']['id']}")
        print(f"Customer: {result['project']['customer']}")
        print(f"Number of product groups: {len(result['product_groups'])}")
        print(f"Grand total: {result['totals']['grand_total']}")
        
    except Exception as e:
        logger.error(f"Error parsing Excel file: {e}")
        raise 