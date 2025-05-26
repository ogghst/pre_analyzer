"""
Analisi Profittabilita Excel to JSON Parser
Converts Excel files with profitability analysis format to structured JSON according to schema
"""

import json
import logging
from typing import Dict, List, Optional, Any
import pandas as pd
from openpyxl import load_workbook

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# =============================================================================
# CONSTANTS
# =============================================================================

# Excel Column Indices (1-based) - Complete mapping for all 81 columns
class ExcelColumns:
    # Basic project columns (1-21)
    COD = 1                          # COD
    PRIORITY_ORDER = 2               # Priorità per ordinamento
    PRIORITY = 3                     # Priorità
    LINE_NUMBER = 4                  # NUMERA LE RIGHE
    WBS = 5                          # WBS
    WBE = 6                          # WBE
    POSITION = 7                     # POS.
    CODICE = 8                       # CODICE
    COD_LISTINO = 9                  # COD.LISTINO
    DENOMINAZIONE = 10               # DENOMINAZIONE
    QTA = 11                         # QTA'
    SUB_TOT_LISTINO = 12             # SUB TOT. LISTINO
    LIST_UNIT = 13                   # LIST.UNIT.
    LISTINO_TOTALE = 14              # LISTINO TOTALE
    SUBTOT_COSTO = 15                # SUBTOT COSTO
    COSTO_UNITARIO = 16              # COSTO UNITARIO
    COSTO_TOTALE = 17                # COSTO TOTALE
    # Column 18 is empty
    COD_2 = 19                       # COD (second occurrence)
    # Column 20 is empty
    TOTALE = 21                      # TOTALE
    
    # Material and UTM columns (22-30)
    MAT = 22                         # MAT
    UTM_ROBOT = 23                   # UTM-ROBOT
    UTM_ROBOT_H = 24                 # UTM-ROBOT-H
    UTM_LGV = 25                     # UTM-LGV
    UTM_LGV_H = 26                   # UTM-LGV-H
    UTM_INTRA = 27                   # UTM-INTRA
    UTM_INTRA_H = 28                 # UTM-INTRA-H
    UTM_LAYOUT = 29                  # UTM-LAYOUT
    UTM_LAYOUT_H = 30                # UTM-LAYOUT-H
    
    # Engineering columns (31-40)
    UTE = 31                         # UTE
    UTE_H = 32                       # UTE-H
    BA = 33                          # BA
    BA_H = 34                        # BA-H
    SW_PC = 35                       # SW-PC
    SW_PC_H = 36                     # SW-PC-H
    SW_PLC = 37                      # SW-PLC
    SW_PLC_H = 38                    # SW-PLC-H
    SW_LGV = 39                      # SW-LGV
    SW_LGV_H = 40                    # SW-LGV-H
    
    # Manufacturing columns (41-50)
    MTG_MEC = 41                     # MTG-MEC
    MTG_MEC_H = 42                   # MTG-MEC-H
    MTG_MEC_INTRA = 43               # MTG-MEC-INTRA
    MTG_MEC_INTRA_H = 44             # MTG-MEC-INTRA-H
    CAB_ELE = 45                     # CAB-ELE
    CAB_ELE_H = 46                   # CAB-ELE-H
    CAB_ELE_INTRA = 47               # CAB-ELE-INTRA
    CAB_ELE_INTRA_H = 48             # CAB-ELE-INTRA-H
    COLL_BA = 49                     # COLL-BA
    COLL_BA_H = 50                   # COLL-BA-H
    
    # Testing columns (51-60)
    COLL_PC = 51                     # COLL-PC
    COLL_PC_H = 52                   # COLL-PC-H
    COLL_PLC = 53                    # COLL-PLC
    COLL_PLC_H = 54                  # COLL-PLC-H
    COLL_LGV = 55                    # COLL-LGV
    COLL_LGV_H = 56                  # COLL-LGV-H
    PM_COST = 57                     # PM
    PM_H = 58                        # PM-H
    SPESE_PM = 59                    # SPESE-PM
    DOCUMENT = 60                    # DOCUMENT
    
    # Logistics and field columns (61-70)
    DOCUMENT_H = 61                  # DOCUMENT-H
    IMBALLO = 62                     # IMBALLO
    STOCCAGGIO = 63                  # STOCCAGGIO
    TRASPORTO = 64                   # TRASPORTO
    SITE = 65                        # SITE
    SITE_H = 66                      # SITE-H
    INSTALL = 67                     # INSTALL
    INSTALL_H = 68                   # INSTALL-H
    AV_PC = 69                       # AV-PC
    AV_PC_H = 70                     # AV-PC-H
    
    # Additional field columns (71-81)
    AV_PLC = 71                      # AV-PLC
    AV_PLC_H = 72                    # AV-PLC-H
    AV_LGV = 73                      # AV-LGV
    AV_LGV_H = 74                    # AV-LGV-H
    SPESE_FIELD = 75                 # SPESE FIELD
    SPESE_VARIE = 76                 # SPESE VARIE
    AFTER_SALES = 77                 # AFTER SALES
    PROVVIGIONI_ITALIA = 78          # PROVVIGIONI ITALIA
    PROVVIGIONI_ESTERO = 79          # PROVVIGIONI ESTERO
    TESORETTO = 80                   # TESORETTO
    MONTAGGIO_BEMA_MBE_US = 81       # MONTAGGIO BEMA MBE-US

# Excel Row Constants
class ExcelRows:
    HEADER_ROW = 3
    DATA_START_ROW = 4
    PROJECT_INFO_START = 1
    PROJECT_INFO_END = 6

# Identification Patterns
class IdentificationPatterns:
    GROUP_PREFIX = 'TXT'
    CATEGORY_CODE_LENGTH = 4
    HEADER_CODE = 'COD'

# Project Information Cell Positions (row, column)
class ProjectInfoCells:
    PROJECT_ID = (1, 1)        # Row 1, Column A - Project identifier
    LISTINO = (2, 1)           # Row 2, Column A - Price list info

# JSON Field Names - Extended to include all columns
class JsonFields:
    # Project fields
    PROJECT = "project"
    ID = "id"
    LISTINO = "listino"
    PARAMETERS = "parameters"
    SALES_INFO = "sales_info"
    
    # Parameter fields (if available in future versions)
    DOC_PERCENTAGE = "doc_percentage"
    PM_PERCENTAGE = "pm_percentage"
    FINANCIAL_COSTS = "financial_costs"
    CURRENCY = "currency"
    EXCHANGE_RATE = "exchange_rate"
    WASTE_DISPOSAL = "waste_disposal"
    WARRANTY_PERCENTAGE = "warranty_percentage"
    IS_24H_SERVICE = "is_24h_service"
    
    # Sales info fields
    AREA_MANAGER = "area_manager"
    AGENT = "agent"
    COMMISSION_PERCENTAGE = "commission_percentage"
    AUTHOR = "author"
    
    # Product group fields
    PRODUCT_GROUPS = "product_groups"
    GROUP_ID = "group_id"
    GROUP_NAME = "group_name"
    QUANTITY = "quantity"
    CATEGORIES = "categories"
    
    # Category fields
    CATEGORY_ID = "category_id"
    CATEGORY_NAME = "category_name"
    CATEGORY_CODE = "category_code"
    WBE = "wbe"
    ITEMS = "items"
    PRICELIST_SUBTOTAL = "pricelist_subtotal"
    COST_SUBTOTAL = "cost_subtotal"
    TOTAL_COST = "total_cost"
    
    # Basic item fields
    POSITION = "position"
    CODE = "code"
    COD_LISTINO = "cod_listino"
    DESCRIPTION = "description"
    QTY = "quantity"
    PRICELIST_UNIT_PRICE = "list_unit_price"
    PRICELIST_TOTAL = "listino_total"
    UNIT_COST = "unit_cost"
    TOTAL_COST = "total_cost"
    INTERNAL_CODE = "internal_code"
    PRIORITY_ORDER = "priority_order"
    PRIORITY = "priority"
    LINE_NUMBER = "line_number"
    WBS = "wbs"
    TOTAL = "totale"
    
    # Material and UTM fields
    MAT = "mat"
    UTM_ROBOT = "utm_robot"
    UTM_ROBOT_H = "utm_robot_h"
    UTM_LGV = "utm_lgv"
    UTM_LGV_H = "utm_lgv_h"
    UTM_INTRA = "utm_intra"
    UTM_INTRA_H = "utm_intra_h"
    UTM_LAYOUT = "utm_layout"
    UTM_LAYOUT_H = "utm_layout_h"
    
    # Engineering fields
    UTE = "ute"
    UTE_H = "ute_h"
    BA = "ba"
    BA_H = "ba_h"
    SW_PC = "sw_pc"
    SW_PC_H = "sw_pc_h"
    SW_PLC = "sw_plc"
    SW_PLC_H = "sw_plc_h"
    SW_LGV = "sw_lgv"
    SW_LGV_H = "sw_lgv_h"
    
    # Manufacturing fields
    MTG_MEC = "mtg_mec"
    MTG_MEC_H = "mtg_mec_h"
    MTG_MEC_INTRA = "mtg_mec_intra"
    MTG_MEC_INTRA_H = "mtg_mec_intra_h"
    CAB_ELE = "cab_ele"
    CAB_ELE_H = "cab_ele_h"
    CAB_ELE_INTRA = "cab_ele_intra"
    CAB_ELE_INTRA_H = "cab_ele_intra_h"
    COLL_BA = "coll_ba"
    COLL_BA_H = "coll_ba_h"
    
    # Testing fields
    COLL_PC = "coll_pc"
    COLL_PC_H = "coll_pc_h"
    COLL_PLC = "coll_plc"
    COLL_PLC_H = "coll_plc_h"
    COLL_LGV = "coll_lgv"
    COLL_LGV_H = "coll_lgv_h"
    PM_COST = "pm_cost"
    PM_H = "pm_h"
    SPESE_PM = "spese_pm"
    DOCUMENT = "document"
    
    # Logistics and field fields
    DOCUMENT_H = "document_h"
    IMBALLO = "imballo"
    STOCCAGGIO = "stoccaggio"
    TRASPORTO = "trasporto"
    SITE = "site"
    SITE_H = "site_h"
    INSTALL = "install"
    INSTALL_H = "install_h"
    AV_PC = "av_pc"
    AV_PC_H = "av_pc_h"
    
    # Additional field fields
    AV_PLC = "av_plc"
    AV_PLC_H = "av_plc_h"
    AV_LGV = "av_lgv"
    AV_LGV_H = "av_lgv_h"
    SPESE_FIELD = "spese_field"
    SPESE_VARIE = "spese_varie"
    AFTER_SALES = "after_sales"
    PROVVIGIONI_ITALIA = "provvigioni_italia"
    PROVVIGIONI_ESTERO = "provvigioni_estero"
    TESORETTO = "tesoretto"
    MONTAGGIO_BEMA_MBE_US = "montaggio_bema_mbe_us"
    
    # Totals fields
    TOTALS = "totals"
    EQUIPMENT_TOTAL = "equipment_total"
    INSTALLATION_TOTAL = "installation_total"
    SUBTOTAL = "subtotal"
    TOTAL_LISTINO = "total_listino"
    TOTAL_COSTO = "total_costo"
    MARGIN = "margin"
    MARGIN_PERCENTAGE = "margin_percentage"

# Calculation Constants
class CalculationConstants:
    DEFAULT_QUANTITY = 1
    DEFAULT_FLOAT = 0.0
    DEFAULT_INT = 0

# Log Messages
class LogMessages:
    PARSING_START = "Starting to parse {}"
    WORKBOOK_LOADED = "Loaded workbook with {} rows and {} columns"
    GROUP_FOUND = "Found group: {}"
    CATEGORY_FOUND = "Found category: {}"
    ITEM_FOUND = "Found item: {}"
    PARSING_COMPLETED = "Parsing completed. Found {} product groups"
    JSON_SAVED = "JSON output saved to {}"

# Error Messages
class ErrorMessages:
    FILE_NOT_FOUND = "File not found: {}"
    INVALID_WORKBOOK = "Unable to load workbook: {}"
    MISSING_SHEET = "Sheet '{}' not found in workbook"
    INVALID_DATA_FORMAT = "Invalid data format in row {}"

# =============================================================================
# MAIN PARSER CLASS
# =============================================================================

class AnalisiProfittabilitaParser:
    """Parser for converting Analisi Profittabilita Excel files to JSON format"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.ws = None
        
    def load_workbook(self):
        """Load the Excel workbook"""
        try:
            self.workbook = load_workbook(self.file_path, data_only=True)
            # Use the first worksheet (typically 'NEW_OFFER1')
            self.ws = self.workbook['NEW_OFFER1']
            logger.info(LogMessages.WORKBOOK_LOADED.format(self.ws.max_row, self.ws.max_column))
        except FileNotFoundError:
            raise FileNotFoundError(ErrorMessages.FILE_NOT_FOUND.format(self.file_path))
        except Exception as e:
            raise Exception(ErrorMessages.INVALID_WORKBOOK.format(str(e)))
    
    def extract_project_info(self) -> Dict[str, Any]:
        """Extract project information from the Excel file"""
        # Extract basic project info
        project_id_cell = self.ws.cell(row=ProjectInfoCells.PROJECT_ID[0], column=ProjectInfoCells.PROJECT_ID[1])
        listino_cell = self.ws.cell(row=ProjectInfoCells.LISTINO[0], column=ProjectInfoCells.LISTINO[1])
        
        project_id = str(project_id_cell.value) if project_id_cell.value else ""
        listino = str(listino_cell.value) if listino_cell.value else ""
        
        return {
            JsonFields.ID: project_id,
            JsonFields.LISTINO: listino,
            JsonFields.PARAMETERS: {
                JsonFields.DOC_PERCENTAGE: CalculationConstants.DEFAULT_FLOAT,
                JsonFields.PM_PERCENTAGE: CalculationConstants.DEFAULT_FLOAT,
                JsonFields.FINANCIAL_COSTS: CalculationConstants.DEFAULT_FLOAT,
                JsonFields.CURRENCY: "EUR",  # Default currency
                JsonFields.EXCHANGE_RATE: 1.0,
                JsonFields.WASTE_DISPOSAL: CalculationConstants.DEFAULT_FLOAT,
                JsonFields.WARRANTY_PERCENTAGE: CalculationConstants.DEFAULT_FLOAT,
                JsonFields.IS_24H_SERVICE: False
            },
            JsonFields.SALES_INFO: {
                JsonFields.AREA_MANAGER: None,
                JsonFields.AGENT: None,
                JsonFields.COMMISSION_PERCENTAGE: CalculationConstants.DEFAULT_FLOAT,
                JsonFields.AUTHOR: None
            }
        }
    
    def _safe_cell_value(self, row: int, column: int, default: Any = None) -> Any:
        """
        Safely extract cell value by row and column index
        
        Args:
            row: Row number (1-based)
            column: Column number (1-based)
            default: Default value if cell is empty or error occurs
            
        Returns:
            Cell value or default
        """
        try:
            cell_value = self.ws.cell(row=row, column=column).value
            if cell_value is None:
                return default
            return cell_value
        except Exception:
            return default
    
    def _safe_cell_float(self, row: int, column: int, default: float = 0.0) -> float:
        """Safely extract float value from cell"""
        return self._safe_float(self._safe_cell_value(row, column), default)
    
    def _safe_cell_int(self, row: int, column: int, default: int = 0) -> int:
        """Safely extract integer value from cell"""
        return self._safe_int(self._safe_cell_value(row, column), default)
    
    def _safe_cell_str(self, row: int, column: int, default: str = "") -> str:
        """Safely extract string value from cell"""
        value = self._safe_cell_value(row, column, default)
        return str(value) if value is not None else default
    
    def extract_product_groups(self) -> List[Dict[str, Any]]:
        """Extract product groups, categories, and items with all columns using safe column access"""
        product_groups = []
        current_group = None
        current_category = None
        
        # Start from data start row
        for row in range(ExcelRows.DATA_START_ROW, self.ws.max_row + 1):
            
            # Skip row if no priority value
            if not self._safe_cell_value(row, ExcelColumns.PRIORITY):
                continue
            
            # Extract basic identification values using safe column access
            cod_val = self._safe_cell_value(row, ExcelColumns.COD)
            codice_val = self._safe_cell_value(row, ExcelColumns.CODICE)
            denominazione_val = self._safe_cell_value(row, ExcelColumns.DENOMINAZIONE)
            qta_val = self._safe_cell_value(row, ExcelColumns.QTA)
            wbe_val = self._safe_cell_value(row, ExcelColumns.WBE)

            # Check if this is a group header (TXT in CODICE)
            if codice_val and str(codice_val).startswith(IdentificationPatterns.GROUP_PREFIX):
                # Save previous group if exists
                if current_group:
                    product_groups.append(current_group)
                
                # Start new group
                current_group = {
                    JsonFields.GROUP_ID: str(codice_val),
                    JsonFields.GROUP_NAME: str(denominazione_val) if denominazione_val else "",
                    JsonFields.QUANTITY: self._safe_int(qta_val, CalculationConstants.DEFAULT_QUANTITY),
                    JsonFields.CATEGORIES: []
                }
                current_category = None
                logger.info(LogMessages.GROUP_FOUND.format(codice_val))
                
            # Check if this is a category (4-char code in COD column)
            elif cod_val and len(str(cod_val).strip()) == IdentificationPatterns.CATEGORY_CODE_LENGTH and current_group:
                current_category = {
                    JsonFields.CATEGORY_ID: str(cod_val),
                    JsonFields.CATEGORY_CODE: str(codice_val) if codice_val else "",
                    JsonFields.CATEGORY_NAME: str(denominazione_val) if denominazione_val else "",
                    JsonFields.WBE: str(wbe_val) if wbe_val else "",
                    JsonFields.PRICELIST_SUBTOTAL: self._safe_cell_float(row, ExcelColumns.SUB_TOT_LISTINO),
                    JsonFields.COST_SUBTOTAL: self._safe_cell_float(row, ExcelColumns.SUBTOT_COSTO),
                    JsonFields.TOTAL_COST: self._safe_cell_float(row, ExcelColumns.COSTO_TOTALE),
                    JsonFields.ITEMS: [],
                }
                current_group[JsonFields.CATEGORIES].append(current_category)
                logger.info(LogMessages.CATEGORY_FOUND.format(cod_val))
                
            # Check if this is an item
            elif denominazione_val and current_category \
                and not (codice_val and str(codice_val).startswith(IdentificationPatterns.GROUP_PREFIX)) \
                and not (cod_val and len(str(cod_val).strip()) == IdentificationPatterns.CATEGORY_CODE_LENGTH) \
                and str(denominazione_val) != "DENOMINAZIONE":  # Skip header row
                
                item = {
                    # Basic identification - using safe column access
                    JsonFields.POSITION: str(row),
                    JsonFields.CODE: self._safe_cell_str(row, ExcelColumns.CODICE),
                    JsonFields.COD_LISTINO: self._safe_cell_str(row, ExcelColumns.COD_LISTINO),
                    JsonFields.DESCRIPTION: str(denominazione_val),
                    JsonFields.QTY: self._safe_cell_float(row, ExcelColumns.QTA),
                    JsonFields.PRICELIST_UNIT_PRICE: self._safe_cell_float(row, ExcelColumns.LIST_UNIT),
                    JsonFields.PRICELIST_TOTAL: self._safe_cell_float(row, ExcelColumns.LISTINO_TOTALE),
                    JsonFields.UNIT_COST: self._safe_cell_float(row, ExcelColumns.COSTO_UNITARIO),
                    JsonFields.TOTAL_COST: self._safe_cell_float(row, ExcelColumns.COSTO_TOTALE),
                    JsonFields.INTERNAL_CODE: self._safe_cell_str(row, ExcelColumns.COD_2),
                    JsonFields.PRIORITY_ORDER: self._safe_cell_int(row, ExcelColumns.PRIORITY_ORDER),
                    JsonFields.PRIORITY: self._safe_cell_int(row, ExcelColumns.PRIORITY),
                    JsonFields.LINE_NUMBER: self._safe_cell_int(row, ExcelColumns.LINE_NUMBER),
                    JsonFields.WBS: self._safe_cell_str(row, ExcelColumns.WBS),
                    JsonFields.TOTAL: self._safe_cell_float(row, ExcelColumns.TOTALE),
                    
                    # Material and UTM fields - using safe column access
                    JsonFields.MAT: self._safe_cell_float(row, ExcelColumns.MAT),
                    JsonFields.UTM_ROBOT: self._safe_cell_float(row, ExcelColumns.UTM_ROBOT),
                    JsonFields.UTM_ROBOT_H: self._safe_cell_float(row, ExcelColumns.UTM_ROBOT_H),
                    JsonFields.UTM_LGV: self._safe_cell_float(row, ExcelColumns.UTM_LGV),
                    JsonFields.UTM_LGV_H: self._safe_cell_float(row, ExcelColumns.UTM_LGV_H),
                    JsonFields.UTM_INTRA: self._safe_cell_float(row, ExcelColumns.UTM_INTRA),
                    JsonFields.UTM_INTRA_H: self._safe_cell_float(row, ExcelColumns.UTM_INTRA_H),
                    JsonFields.UTM_LAYOUT: self._safe_cell_float(row, ExcelColumns.UTM_LAYOUT),
                    JsonFields.UTM_LAYOUT_H: self._safe_cell_float(row, ExcelColumns.UTM_LAYOUT_H),
                    
                    # Engineering fields - using safe column access
                    JsonFields.UTE: self._safe_cell_float(row, ExcelColumns.UTE),
                    JsonFields.UTE_H: self._safe_cell_float(row, ExcelColumns.UTE_H),
                    JsonFields.BA: self._safe_cell_float(row, ExcelColumns.BA),
                    JsonFields.BA_H: self._safe_cell_float(row, ExcelColumns.BA_H),
                    JsonFields.SW_PC: self._safe_cell_float(row, ExcelColumns.SW_PC),
                    JsonFields.SW_PC_H: self._safe_cell_float(row, ExcelColumns.SW_PC_H),
                    JsonFields.SW_PLC: self._safe_cell_float(row, ExcelColumns.SW_PLC),
                    JsonFields.SW_PLC_H: self._safe_cell_float(row, ExcelColumns.SW_PLC_H),
                    JsonFields.SW_LGV: self._safe_cell_float(row, ExcelColumns.SW_LGV),
                    JsonFields.SW_LGV_H: self._safe_cell_float(row, ExcelColumns.SW_LGV_H),
                    
                    # Manufacturing fields - using safe column access
                    JsonFields.MTG_MEC: self._safe_cell_float(row, ExcelColumns.MTG_MEC),
                    JsonFields.MTG_MEC_H: self._safe_cell_float(row, ExcelColumns.MTG_MEC_H),
                    JsonFields.MTG_MEC_INTRA: self._safe_cell_float(row, ExcelColumns.MTG_MEC_INTRA),
                    JsonFields.MTG_MEC_INTRA_H: self._safe_cell_float(row, ExcelColumns.MTG_MEC_INTRA_H),
                    JsonFields.CAB_ELE: self._safe_cell_float(row, ExcelColumns.CAB_ELE),
                    JsonFields.CAB_ELE_H: self._safe_cell_float(row, ExcelColumns.CAB_ELE_H),
                    JsonFields.CAB_ELE_INTRA: self._safe_cell_float(row, ExcelColumns.CAB_ELE_INTRA),
                    JsonFields.CAB_ELE_INTRA_H: self._safe_cell_float(row, ExcelColumns.CAB_ELE_INTRA_H),
                    JsonFields.COLL_BA: self._safe_cell_float(row, ExcelColumns.COLL_BA),
                    JsonFields.COLL_BA_H: self._safe_cell_float(row, ExcelColumns.COLL_BA_H),
                    
                    # Testing fields - using safe column access
                    JsonFields.COLL_PC: self._safe_cell_float(row, ExcelColumns.COLL_PC),
                    JsonFields.COLL_PC_H: self._safe_cell_float(row, ExcelColumns.COLL_PC_H),
                    JsonFields.COLL_PLC: self._safe_cell_float(row, ExcelColumns.COLL_PLC),
                    JsonFields.COLL_PLC_H: self._safe_cell_float(row, ExcelColumns.COLL_PLC_H),
                    JsonFields.COLL_LGV: self._safe_cell_float(row, ExcelColumns.COLL_LGV),
                    JsonFields.COLL_LGV_H: self._safe_cell_float(row, ExcelColumns.COLL_LGV_H),
                    JsonFields.PM_COST: self._safe_cell_float(row, ExcelColumns.PM_COST),
                    JsonFields.PM_H: self._safe_cell_float(row, ExcelColumns.PM_H),
                    JsonFields.SPESE_PM: self._safe_cell_float(row, ExcelColumns.SPESE_PM),
                    JsonFields.DOCUMENT: self._safe_cell_float(row, ExcelColumns.DOCUMENT),
                    
                    # Logistics and field fields - using safe column access
                    JsonFields.DOCUMENT_H: self._safe_cell_float(row, ExcelColumns.DOCUMENT_H),
                    JsonFields.IMBALLO: self._safe_cell_float(row, ExcelColumns.IMBALLO),
                    JsonFields.STOCCAGGIO: self._safe_cell_float(row, ExcelColumns.STOCCAGGIO),
                    JsonFields.TRASPORTO: self._safe_cell_float(row, ExcelColumns.TRASPORTO),
                    JsonFields.SITE: self._safe_cell_float(row, ExcelColumns.SITE),
                    JsonFields.SITE_H: self._safe_cell_float(row, ExcelColumns.SITE_H),
                    JsonFields.INSTALL: self._safe_cell_float(row, ExcelColumns.INSTALL),
                    JsonFields.INSTALL_H: self._safe_cell_float(row, ExcelColumns.INSTALL_H),
                    JsonFields.AV_PC: self._safe_cell_float(row, ExcelColumns.AV_PC),
                    JsonFields.AV_PC_H: self._safe_cell_float(row, ExcelColumns.AV_PC_H),
                    
                    # Additional field fields - using safe column access
                    JsonFields.AV_PLC: self._safe_cell_float(row, ExcelColumns.AV_PLC),
                    JsonFields.AV_PLC_H: self._safe_cell_float(row, ExcelColumns.AV_PLC_H),
                    JsonFields.AV_LGV: self._safe_cell_float(row, ExcelColumns.AV_LGV),
                    JsonFields.AV_LGV_H: self._safe_cell_float(row, ExcelColumns.AV_LGV_H),
                    JsonFields.SPESE_FIELD: self._safe_cell_float(row, ExcelColumns.SPESE_FIELD),
                    JsonFields.SPESE_VARIE: self._safe_cell_float(row, ExcelColumns.SPESE_VARIE),
                    JsonFields.AFTER_SALES: self._safe_cell_float(row, ExcelColumns.AFTER_SALES),
                    JsonFields.PROVVIGIONI_ITALIA: self._safe_cell_float(row, ExcelColumns.PROVVIGIONI_ITALIA),
                    JsonFields.PROVVIGIONI_ESTERO: self._safe_cell_float(row, ExcelColumns.PROVVIGIONI_ESTERO),
                    JsonFields.TESORETTO: self._safe_cell_float(row, ExcelColumns.TESORETTO),
                    JsonFields.MONTAGGIO_BEMA_MBE_US: self._safe_cell_float(row, ExcelColumns.MONTAGGIO_BEMA_MBE_US)
                }
                
                current_category[JsonFields.ITEMS].append(item)
                logger.debug(LogMessages.ITEM_FOUND.format(codice_val))
        
        # Add the last group if exists
        if current_group:
            product_groups.append(current_group)
        
        # Calculate category totals
        for group in product_groups:
            for category in group[JsonFields.CATEGORIES]:
                if not category[JsonFields.PRICELIST_SUBTOTAL]:
                    category[JsonFields.PRICELIST_SUBTOTAL] = sum(
                        item[JsonFields.PRICELIST_TOTAL] for item in category[JsonFields.ITEMS]
                    )
                if not category[JsonFields.COST_SUBTOTAL]:
                    category[JsonFields.COST_SUBTOTAL] = sum(
                        item[JsonFields.TOTAL_COST] for item in category[JsonFields.ITEMS]
                    )
                if not category[JsonFields.TOTAL_COST]:
                    category[JsonFields.TOTAL_COST] = category[JsonFields.COST_SUBTOTAL]
        
        return product_groups
    
    def calculate_totals(self, product_groups: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Calculate total costs, revenues and margins"""
        total_listino = CalculationConstants.DEFAULT_FLOAT
        total_costo = CalculationConstants.DEFAULT_FLOAT
        
        # Sum up all costs and revenues from all categories
        for group in product_groups:
            for category in group[JsonFields.CATEGORIES]:
                total_listino += category.get(JsonFields.PRICELIST_SUBTOTAL, 0)
                total_costo += category.get(JsonFields.COST_SUBTOTAL, 0)
        
        # Calculate margin
        margin = total_listino - total_costo
        margin_percentage = (margin / total_listino * 100) if total_listino > 0 else 0
        
        return {
            JsonFields.TOTAL_LISTINO: round(total_listino, 2),
            JsonFields.TOTAL_COSTO: round(total_costo, 2),
            JsonFields.MARGIN: round(margin, 2),
            JsonFields.MARGIN_PERCENTAGE: round(margin_percentage, 2),
            JsonFields.EQUIPMENT_TOTAL: round(total_listino, 2),  # For compatibility
            JsonFields.INSTALLATION_TOTAL: CalculationConstants.DEFAULT_FLOAT,
            JsonFields.SUBTOTAL: round(total_listino, 2)
        }
    
    def parse(self) -> Dict[str, Any]:
        """Main parsing method"""
        logger.info(LogMessages.PARSING_START.format(self.file_path))
        
        self.load_workbook()
        
        # Extract all sections
        project_info = self.extract_project_info()
        product_groups = self.extract_product_groups()
        totals = self.calculate_totals(product_groups)
        
        # Build final structure
        result = {
            JsonFields.PROJECT: project_info,
            JsonFields.PRODUCT_GROUPS: product_groups,
            JsonFields.TOTALS: totals
        }
        
        logger.info(LogMessages.PARSING_COMPLETED.format(len(product_groups)))
        return result
    
    def _safe_float(self, value: Any, default: float = CalculationConstants.DEFAULT_FLOAT) -> float:
        """Safely convert value to float"""
        if value is None:
            return default
        try:
            return float(value)
        except (ValueError, TypeError):
            return default
    
    def _safe_int(self, value: Any, default: int = CalculationConstants.DEFAULT_INT) -> int:
        """Safely convert value to int"""
        if value is None:
            return default
        try:
            return int(value)
        except (ValueError, TypeError):
            return default

def parse_analisi_profittabilita_to_json(file_path: str, output_path: Optional[str] = None) -> Dict[str, Any]:
    """
    Main function to parse Analisi Profittabilita Excel file and optionally save to JSON
    
    Args:
        file_path: Path to the Excel file
        output_path: Optional path to save JSON output
        
    Returns:
        Dictionary containing the parsed data
    """
    parser = AnalisiProfittabilitaParser(file_path)
    result = parser.parse()
    
    if output_path:
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, indent=2, ensure_ascii=False)
        logger.info(LogMessages.JSON_SAVED.format(output_path))
    
    return result


if __name__ == "__main__":
    # Example usage
    input_file = "input/cc2199_analisi_profittabilita_new_offer1.xlsx"
    output_file = "output/analisi_profittabilita_complete.json"
    
    try:
        result = parse_analisi_profittabilita_to_json(input_file, output_file)
        print("Parsing completed successfully!")
        print(f"Project ID: {result['project']['id']}")
        print(f"Listino: {result['project']['listino']}")
        print(f"Number of product groups: {len(result['product_groups'])}")
        print(f"Total Listino: {result['totals']['total_listino']}")
        print(f"Total Costo: {result['totals']['total_costo']}")
        print(f"Margin: {result['totals']['margin']} ({result['totals']['margin_percentage']}%)")
        
        # Show sample of additional fields
        if result['product_groups']:
            first_group = result['product_groups'][0]
            if first_group['categories']:
                first_category = first_group['categories'][0]
                if first_category['items']:
                    sample_item = first_category['items'][0]
                    print(f"\nSample additional fields from first item:")
                    print(f"  UTM Robot: {sample_item.get('utm_robot', 0)}")
                    print(f"  PM Cost: {sample_item.get('pm_cost', 0)}")
                    print(f"  Install: {sample_item.get('install', 0)}")
                    print(f"  After Sales: {sample_item.get('after_sales', 0)}")
        
    except Exception as e:
        logger.error(f"Error parsing Excel file: {e}")
        raise 